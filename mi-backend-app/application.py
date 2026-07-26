
import json
import pymysql
import os
from io import BytesIO
import boto3
import dropbox
import uuid
from datetime import datetime,date
import botocore
import base64
from flask import Response
from jinja2 import Template
from urllib.parse import quote
from pathlib import PurePosixPath
import mimetypes
import hashlib
from flask import render_template_string
from sqlalchemy import func
from types import SimpleNamespace
from unidecode import unidecode
import requests

from itsdangerous import URLSafeTimedSerializer

from openpyxl.styles import Font, PatternFill


import pandas as pd



from sqlalchemy.orm import aliased


import re


import zipfile


from bs4 import BeautifulSoup

from flask import (flash, jsonify, make_response, redirect, render_template,
                   request, session, url_for)
from flask_babel import Babel, _
from flask_jwt_extended import (create_access_token, get_jwt_identity,
                                jwt_required, set_access_cookies)
from sqlalchemy import and_, asc, case
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import joinedload






from app_init import bcrypt, create_app, db
from creacion_BD import crear_base_si_no_existe
from funciones import (create_reset_token,lead_exists_for_prospect,
                       send_new_password, update_user_password,
                       validate_reset_token, get_dropbox_access_token,subir_a_dropbox)

from models import (db, Campaign, Newsletter, User,CampaignRecipient, LeadForm, LeadTarget,LeadTargetItem,LeadCampaignHistory,ProspectsIA,  
                            ProspectTarget, ProspectTargetItem)


from config import (BD ,EMAIL_USER,EMAIL_PASSWORD,URL_CONTACTO ,URL_OFERTAS,URL_ACTUALIZAR_CONTACTO,
                     API_KEY,ENVIRONMENT,SEND_EMAIL,SEND_WELLCOME_EMAIL,URL_PROFORMAS,URL_FORM_CONTACTO,
                    AWS_REGION,S3_BUCKET,ROOT_PREFIX_S3,ROOT_PREFIX_DROPBOX, BASE_URL)


from funciones_generar_email import (build_framework,slugify,
                                     extract_html_inline_and_attachments_from_eml_bytes,
                                     rehost_images_under_template_from_html,
                                     resolve_cid_with_attachments,get_db_credentials,
                                     insert_extra_files_into_html,
                                     extract_default_context_from_html,
                                     replace_cid_srcs_with_urls,
                                     fix_relative_imgs,
                                     replace_cid_everywhere,clean_signature_images,
                                     inject_preview_css,put_public_s3,
                                     public_url, parent_of,normalize_incoming_content,
                                     update_manifest,update_manifest_for_key,
                                     apply_manifest_images_all,unescape_pre_wrapped_html,
                                     manifest_lookup,_attachments_html,
                                     enforce_dimensions_from_manifest,
                                    insert_extra_files_into_html, _inner_body_html,
                                    s3_key_exists,_norm_src,_collect_image_keys,split_body_and_signature,
                                    _coerce_items,paths, TEMPLATES_ROOT, get_s3,
                                    _upload_zip_assets_and_rewrite_html,_guess_content_type_from_name,
                                    _find_html_in_zip,send_email_ses,load_newsletter_html,
                                    build_final_email_html,send_campaign_batch,send_campaign_batch_stream,
                                    USE_S3, S3_BUCKET,key_message, key_original,key_template, key_signature,
                                    s3_get_text, s3_put_text,BASE_DIR)
                                     


from flask import Response, stream_with_context

application = create_app()

@application.template_filter('escapejs')
def escapejs_filter(s):
    return json.dumps(str(s))[1:-1]


from decimal import Decimal

@application.route('/')
def index():
    session.pop('_flashes', None)  # Limpia mensajes pendientes manualmente

    return render_template('login.html')






def get_locale():

    lang = session.get('lang') or request.cookies.get('lang') or 'es'
    # Normaliza: solo letras minúsculas (ej. "en-us" -> "en")
    
    if lang:
        # Si el idioma tiene un guion, tomamos solo la parte antes del guion
        # Ejemplo: 'en-us' -> 'en'
        # Esto es útil si se quiere normalizar a un código de idioma más simple
        return lang.split('-')[0].lower()

        # Si no hay idioma en la sesión, lo forzamos a español
    session['lang'] = 'es'
    return 'es'
    


babel = Babel(application, locale_selector=get_locale)


@application.route('/set_language', methods=['POST'])
def set_language():

    lang = request.form.get('language', 'es')
    next_page = request.form.get('next') or url_for('index') 
    # Normaliza si es necesario
    lang = lang.split('-')[0]  # 'en-us' -> 'en'
    session['lang'] = lang
    resp = make_response(redirect(next_page))

    #resp = make_response(redirect(request.referrer or '/'))
    print("🔍 Referrer:", request.referrer) 
    print("resp", resp)
    resp.set_cookie('lang', lang)
    print("set_language", lang)
    return resp

    
   


   

@application.context_processor
def inject_get_locale():
    return dict(get_locale=get_locale)

@application.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        session.pop('_flashes', None)  # Limpia mensajes pendientes manualmente
        data = request.get_json()  # Obtener datos en formato JSON

        email = data.get('email')
        password = data.get('password')

        try:
            user = User.query.filter_by(email=email).first()
        except OperationalError as e:
            db.session.rollback()
            flash(_("Error Operacional."), 'error')
            return jsonify({"msg": _("Error operacional")}), 500

        if user is None:
            flash(_("El email facilitado no existe"), 'error')
            return jsonify({"msg": _("El email facilitado no existe")}), 401
        else:
           
            
            ruta_login = '/main_page'
            
            
            if user and bcrypt.check_password_hash(user.password, password):
                #access_token = create_access_token(identity=str({'email': user.email}))
                access_token = create_access_token(identity=user.email)
                print("Usuario autenticado:", user.email)
                session['email'] = user.email
                session['access_token'] = access_token
                print ("user.uid", user.uid )

                # Creamos primero el diccionario de datos
                response_data = {
                    "msg": _("Login exitoso"),
                    "uid": user.uid,
                    "ruta_login": ruta_login,
                    
                }

                
                # Convertimos el dict a JSON, y luego lo envolvemos con make_response
                json_response = jsonify(response_data)
                response = make_response(json_response)

                # Establecemos la cookie JWT
                set_access_cookies(response, access_token)

                return response

            flash(_("Contraseña incorrecta."), 'error')
            return jsonify({"msg": _("Contraseña incorrecta.")}), 401
    session.pop('_flashes', None)  # Limpia mensajes pendientes manualmente
    return render_template('login.html')
                

        
    

#@app.route('/register', methods=['GET', 'POST'])
#def register():
#    if request.method == 'POST':
#        username = request.form['username']
#        password = request.form['password']
        # Aquí iría la lógica de registro

@application.route('/register', methods=[ 'GET','POST'])
def register():
    
   
    if request.method == 'POST':
        
        data = request.get_json()
       
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        
       

        try :
            user = User.query.filter_by(username=username).first()
        except OperationalError as e:
            db.session.rollback()
            print(f"Error operacional: {e}")
            flash("Error Operacional.", 'error')
            return jsonify({"msg": "Error operacional"}), 500


        if user:
            print (username)
            uid = user.uid

            

            
            


            try :
                #Buscar  cunato clubs tiene es uid
                existing_user_email = User.query.filter( (User.email == email)).first()
            except OperationalError as e:
                db.session.rollback()
                print(f"Error operacional: {e}")
                flash("Error Operacional.", 'error')
                return jsonify({"msg": "Error operacional"}), 500




            # Verificar si el usuario ya existe
            if existing_user_email:  

                print ("existing_user_email", existing_user_email.email)         

                return jsonify({"msg": "La dirección de email ya esá registrada"}), 401
            
            else :    

                print (" NO existing_user_email")
               
                hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')

                
                if not user.email :
                   

                    

                    try :
                        User.query.filter_by(username=username).update({
                        'email': email,
                        'password': hashed_password})

                    
                        print  ("User_email", email)
                        db.session.commit()
               

                    except OperationalError as e:
                        db.session.rollback()
                        print(f"Error operacional: {e}")
                        flash("Error Operacional.", 'error')
                        return jsonify({"msg": "Error operacional"}), 500





                else :

                    print  ("User_email2", email)
                    # Aquí iría la lógica de autenticación
                   
                    try :
                        new_user = User(username=username, email=email, password=hashed_password, uid=uid, uid_hytronik=UID_HYTRONIK)
                        db.session.add(new_user)
                        db.session.commit()


                    except OperationalError as e:
                        db.session.rollback()
                        print(f"Error operacional: {e}")
                        flash("Error Operacional.", 'error')
                        return jsonify({"msg": "Error operacional"}), 500


                    





                flash("Usuario registrado con éxito.", 'success')
                flash("Por favor, inicie sesión.", 'success')
                return redirect(url_for('login'))  # Redirigir al formulario de login


                #return jsonify({"msg": "Usuario registrado correctamente"}), 200
        
        else : 
            
            existe_usuario = db.session.query(User).first()

            if existe_usuario:
                return jsonify({"msg": "El usuario no existe"}), 404
            else:
                print ("admin")
                username = "admin"
                print (email)
                print (password)
                hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')

                
                try :
                        
                    new_user = User(username=username, email=email, password=hashed_password,uid=766)
                    db.session.add(new_user)
                    db.session.commit()
                    flash("Usuario registrado con éxito.", 'success')
                    flash("Por favor, inicie sesión.", 'success')


                except OperationalError as e:
                    db.session.rollback()
                    flash("Error Operacional.", 'error')
                    print(f"Error operacional: {e}")

                return redirect(url_for('login'))  # Redirigir al formulario de login
                
                #return jsonify({"msg": "El usuario admin se ha registrado"}), 200



        
        
        
           
  
    
    return render_template('register.html')





@application.route('/forgot_password', methods=[ 'GET','POST'])


def forgot_password():
    if request.method == 'POST':


        
        #data= { "username": 'aregalado', "password": 'Madrid01' }
        data = request.get_json()  # Obtener datos en formato JSON         
        email = data.get('email')
        

       
       

       

        try :
                        
            user = User.query.filter_by(email=email).first()            

        except OperationalError as e:
            db.session.rollback()
            print(f"Error operacional: {e}")
            flash("Error Operacional.", 'error')
            return jsonify({"msg": "Error operacional"}), 500




        if user :
           

            
            reset_token = create_reset_token(user.id) # Genera un token seguro

            
            # Llamar a la función para enviar el correo con el token
            respuesta = send_new_password(email, reset_token)

            if (respuesta == "error"): 
                    
                return jsonify({"msg":"error al enviar el correo" }), 404
            else :

                flash ("Instrucciones enviadas con exito por email ")
               
                        

                return jsonify({"msg":"Cooreo enviado con éxito" }), 200
        
            


       

        

        
                
            

        return jsonify({"msg": "No existe la dirección de correo"}), 401
    else : 
      
        



        return render_template('Forgot_password.html')





@application.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        token = request.form.get('token')
        new_password = request.form.get('new_password')

        

        # (Aquí deberías validar el token primero)
        if validate_reset_token(token):
            # Si el token es válido, actualizar la contraseña en la base de datos
            hashed_password = bcrypt.generate_password_hash(new_password).decode('utf-8')
            
        
            if update_user_password(token, hashed_password):
                    flash("Contraseña restablecida con éxito.", 'success')
                    flash("Por favor, inicie sesión.", 'success')
                    return redirect(url_for('login'))  # Redirigir al formulario de login
            else:
                    flash("No se pudo actualizar la contraseña.", 'error')
                    return redirect(url_for('reset_password'))  # Redirigir de nuevo a restablecer contraseña
        


        else : 
            

            flash("Tiempo máximo expirado para restablecer la contraseña.", 'error')
            flash("Por favor, vuelva a intentarlo", 'error')
            return redirect(url_for('login'))  # Redirigir al formulario de login
    
    
    
        

    # Si es un GET, mostrar el formulario
    token = request.args.get('token')

    
    return render_template('reset_password.html', token=token)


       





        
               
       #
   
   


@application.route('/main_page', methods=[ 'GET','POST'])

#@jwt_required()



def main_page():

    print("main_page")

    if request.method == "POST":
        data = request.json
        application.logger.info("Datos recibidos: %s", data)
        print("if POST")

    else:

        print("enviando datos del GET main page")

        access_token = session.get("access_token")

        if not access_token:
            return "Token no encontrado en sesión", 401

        response = redirect(
            url_for(
                "consultar_leads",
                estado="Sin calificar",
                modo="leads"
            )
        )

        set_access_cookies(response, access_token)

        return response



@application.route('/ofertas', methods=['GET', 'POST'])
def ofertas():
    if request.method == 'POST':
        from io import BytesIO
        import pycurl
        try:
            data = request.get_json()

            # Recoge los datos enviados desde el frontend
            name = data.get('name')
            email = data.get('email')
            idioma = data.get('idioma')
            pais = data.get('pais')
            tipo_lead = data.get('tipo_lead')
            pistas_perimetrales = data.get('pistas_perimetrales')
            pistas_laterales = data.get('pistas_laterales')
            incluir_transporte = data.get('incluir_transporte', False)
            importe_transporte = data.get('importe_transporte', 0)   
            mailorigen = 'soporte@planetpower.es'
            descuento_adicional = Decimal(data.get("descuento_adicional", 0))
            origen = 'CRM'
            descuentoFijo= Decimal(data.get("descuento_fijo", 0))
 

            print("📥 Datos recibidos:")
            print(f"Nombre: {name}")
            print(f"Email: {email}")
            print(f"Idioma: {idioma}")
            print(f"País: {pais}")
            print(f"Tipo: {tipo_lead}")
            print(f"Descuento adicional: {descuento_adicional}")
            print(f"Pistas perimetrales: {pistas_perimetrales}")
            print(f"Pistas laterales: {pistas_laterales}")
            print(f"Descuento adicional: {descuento_adicional}")
            print(f"Incluir Transporte : {incluir_transporte}")
            print(f"Importe Transporte : {importe_transporte}")
            print ("Send_EMAIL", SEND_EMAIL)
            print ("Send_WellCome_EMAIL", SEND_WELLCOME_EMAIL)
            print(f"Descuento Fijo: {descuentoFijo}")
 




       

            payload = {
                "name": name,
                "email": email,
                "idioma": idioma,
                "pais": pais,
                "tipo_lead": tipo_lead,
                "pistas_perimetrales": pistas_perimetrales,
                "pistas_laterales": pistas_laterales,
                "mailorigen": EMAIL_USER,
                "descuento_adicional": int(descuento_adicional),
                "incluir_transporte": incluir_transporte,
                "importe_transporte": importe_transporte,
                "descuento_fijo": int(descuentoFijo),
                "origen": origen,
                "BD": BD,
                "EMAIL_USER": EMAIL_USER,
                "EMAIL_PASSWORD": EMAIL_PASSWORD,
                "URL_CONTACTO": URL_CONTACTO,
                "URL_OFERTAS": URL_OFERTAS,
                "URL_ACTUALIZAR_CONTACTO": URL_ACTUALIZAR_CONTACTO,
                "URL_FORM_CONTACTO": URL_FORM_CONTACTO,
                "URL_PROFORMAS": URL_PROFORMAS,
                "ENVIRONMENT": ENVIRONMENT,
                "SEND_EMAIL": SEND_EMAIL,
                "SEND_WELLCOME_EMAIL": SEND_WELLCOME_EMAIL,

            }
            


            print ("URL_CONTACTO", URL_CONTACTO)
            print("URL_OFERTAS", URL_OFERTAS)
            print("URL_ACTUALIZAR_CONTACTO", URL_ACTUALIZAR_CONTACTO)
            print("URL_FORM_CONTACTO", URL_FORM_CONTACTO)
            print("URL_PROFORMAS", URL_PROFORMAS)



    
            api_key = API_KEY
          

            api_url = URL_CONTACTO

            headers = [
                "x-api-key: " + api_key,
                "Request-Origin: SwaggerBootstrapUi",
                "Accept: application/json",
                "Content-Type: application/json",
            ]

            body = json.dumps({
                "name": name,
                "email": email,
                "idioma": idioma,
                "pais": pais,
                "tipo_lead": tipo_lead,
                "pistas_perimetrales": pistas_perimetrales,
                "pistas_laterales": pistas_laterales,
                #"mailorigen": mailorigen,
                "mailorigen": EMAIL_USER,
                "descuento_adicional": int (descuento_adicional),
                "incluir_transporte": incluir_transporte,
                "importe_transporte": importe_transporte,
                "descuento_fijo": int(descuentoFijo),
                "origen": origen,
                "BD": BD,
                "EMAIL_USER": EMAIL_USER,
                "EMAIL_PASSWORD": EMAIL_PASSWORD,
                "URL_CONTACTO": URL_CONTACTO,
                "URL_OFERTAS": URL_OFERTAS,
                "URL_ACTUALIZAR_CONTACTO": URL_ACTUALIZAR_CONTACTO,
                "URL_FORMCONTACTO": URL_FORM_CONTACTO,
                "URL_PROFORMAS": URL_PROFORMAS,
                "ENVIRONMENT": ENVIRONMENT ,
                "SEND_EMAIL": SEND_EMAIL

            })


           

            buffer = BytesIO()
            c = pycurl.Curl()
            c.setopt(c.URL, api_url)
            c.setopt(c.POST, 1)
            c.setopt(c.POSTFIELDS, body)
            c.setopt(pycurl.SSL_VERIFYPEER, 0)
            c.setopt(pycurl.SSL_VERIFYHOST, 0)
            c.setopt(pycurl.CONNECTTIMEOUT, 10)
            c.setopt(pycurl.TIMEOUT, 120)
            c.setopt(c.HTTPHEADER, headers)
            c.setopt(c.WRITEDATA, buffer)

            c.perform()
            status_code = c.getinfo(pycurl.RESPONSE_CODE) or 500
            response_body = buffer.getvalue().decode('utf-8')
            c.close()

            print(f"✅ Respuesta del backend (status {status_code}): {response_body}")

            response = make_response(response_body, status_code)
            response.headers["Content-Type"] = "application/json"
            return response

        except Exception as e:
            print("❌ Excepción capturada:", str(e))
            return make_response(jsonify({"error": str(e)}), 500)

    # Si es GET, renderiza el formulario
    return render_template('ofertas.html')




def tipo_identificacion_por_pais_texto(pais_texto: str) -> str:

    if not pais_texto:
        return "REGISTRO"

    p = re.sub(r"\s+", " ", pais_texto).strip()

    try:
        creds = get_db_credentials("secretoBC/Mysql")
        dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

        conn = pymysql.connect(
            host=creds['host'],
            user=creds['username'],
            password=creds['password'],
            database=dbname,
            port=int(creds.get('port', 3306)),
            cursorclass=pymysql.cursors.DictCursor  # 👈 IMPORTANTE
        )

        with conn.cursor() as cur:
            cur.execute("""
                SELECT codigo_pais, mercado
                FROM pais
                WHERE pais_es = %s OR pais_en = %s OR pais_fr = %s OR pais_it = %s
                LIMIT 1
            """, (p, p, p, p))

            row = cur.fetchone()

    except Exception as e:
        print(f"Error consultando paises: {e}")
        return "REGISTRO"

    finally:
        try:
            conn.close()
        except:
            pass

    if not row:
        return "REGISTRO"

    codigo_pais = (row["codigo_pais"] or "").upper()
    mercado = (row["mercado"] or "").upper()

    if codigo_pais == "ES" or mercado == "NACIONAL":
        return "NIF"

    if mercado == "UE":
        return "VAT"

    return "REGISTRO"





@application.route('/facturaProforma', methods=['GET', 'POST'])
def facturaProforma():
    if request.method == 'GET':
        return render_template('ofertas.html')

    # POST
    data = request.get_json(force=True) or {}
    quoteNumber = (data.get('quoteNumber') or '').strip()
    email = (data.get('email') or '').strip()
    idioma = (data.get('idioma') or '').strip()
    pais = (data.get('pais') or '').strip()
    
    lead_id = str(data.get('lead_id') or '').strip()


    print("📥 Datos recibidos:")
    print(f"QuoteNumber: {quoteNumber}")
    print(f"email: {email}")
    print(f"idioma: {idioma}")

    

    # Devolver algo que el front pueda leer como JSON
    return jsonify({
        "ok": True,
        "redirect": f"/contactoFacturaProforma?quoteNumber={quoteNumber}&email={email}&idioma={idioma}&pais={pais}&lead_id={lead_id}"
    })


@application.route('/consultar_campanas_por_prospectoIA')
def consultar_campanas_por_prospectoIA():
    prospectIA_id = request.args.get("prospectIA_id", type=int)

    if not prospectIA_id:
        return jsonify({"error": "missing_prospectIA_id"}), 400

    NewsletterES = aliased(Newsletter)
    NewsletterEN = aliased(Newsletter)

    rows = (
        db.session.query(
            CampaignRecipient.campaign_id,

            Campaign.name.label("campaign_name"),
            Campaign.campaign_type.label("campaign_type"),
            Campaign.idioma.label("campaign_idioma"),

            Campaign.subject_es.label("subject_es"),
            Campaign.subject_en.label("subject_en"),

            Campaign.sent_at.label("campaign_sent_at"),

            CampaignRecipient.sent_at.label("recipient_sent_at"),
            CampaignRecipient.origen.label("origen"),
            CampaignRecipient.idioma.label("recipient_idioma"),
            CampaignRecipient.send_status.label("send_status"),

            CampaignRecipient.opened_at.label("opened_at"),
            CampaignRecipient.clicked_at.label("clicked_at"),
            CampaignRecipient.click_count.label("click_count"),

            NewsletterES.name.label("newsletter_es_name"),
            NewsletterES.template_s3_path.label("newsletter_es_path"),

            NewsletterEN.name.label("newsletter_en_name"),
            NewsletterEN.template_s3_path.label("newsletter_en_path"),
        )
        .join(
            Campaign,
            Campaign.id == CampaignRecipient.campaign_id
        )
        .outerjoin(
            NewsletterES,
            NewsletterES.id == Campaign.newsletter_es_id
        )
        .outerjoin(
            NewsletterEN,
            NewsletterEN.id == Campaign.newsletter_en_id
        )
        .filter(
            CampaignRecipient.entity_kind == "prospect",
            CampaignRecipient.entity_id == prospectIA_id
        )
        .order_by(
            CampaignRecipient.sent_at.desc(),
            CampaignRecipient.campaign_id.desc()
        )
        .all()
    )

    return render_template(
        "campanas_por_prospecto.html",
        rows=rows,
        prospectIA_id=prospectIA_id,
        s3_bucket=S3_BUCKET
    )

@application.route('/consultar_campanas_por_lead')
def consultar_campanas_por_lead():
    lead_id = request.args.get("lead_id", type=int)

    if not lead_id:
        return jsonify({"error": "missing_lead_id"}), 400

    NewsletterES = aliased(Newsletter)
    NewsletterEN = aliased(Newsletter)

    rows = (
        db.session.query(
            CampaignRecipient.campaign_id,

            Campaign.name.label("campaign_name"),
            Campaign.campaign_type.label("campaign_type"),
            Campaign.idioma.label("campaign_idioma"),

            Campaign.subject_es.label("subject_es"),
            Campaign.subject_en.label("subject_en"),

            Campaign.sent_at.label("campaign_sent_at"),

            CampaignRecipient.sent_at.label("recipient_sent_at"),
            CampaignRecipient.origen.label("origen"),
            CampaignRecipient.idioma.label("recipient_idioma"),
            CampaignRecipient.send_status.label("send_status"),

            CampaignRecipient.opened_at.label("opened_at"),
            CampaignRecipient.clicked_at.label("clicked_at"),
            CampaignRecipient.click_count.label("click_count"),

            NewsletterES.name.label("newsletter_es_name"),
            NewsletterES.template_s3_path.label("newsletter_es_path"),

            NewsletterEN.name.label("newsletter_en_name"),
            NewsletterEN.template_s3_path.label("newsletter_en_path"),
        )
        .join(
            Campaign,
            Campaign.id == CampaignRecipient.campaign_id
        )
        .outerjoin(
            NewsletterES,
            NewsletterES.id == Campaign.newsletter_es_id
        )
        .outerjoin(
            NewsletterEN,
            NewsletterEN.id == Campaign.newsletter_en_id
        )
        .filter(
            CampaignRecipient.entity_kind == "lead",
            CampaignRecipient.entity_id == lead_id
        )
        .order_by(
            CampaignRecipient.sent_at.desc(),
            CampaignRecipient.campaign_id.desc()
        )
        .all()
    )

    
    return render_template(
        "campanas_por_lead.html",
        rows=rows,
        lead_id=lead_id,
        s3_bucket=S3_BUCKET
    )


@application.route('/campanas_detalle_por_lead')
def campanas_detalle_por_lead():
    lead_id = request.args.get("lead_id", type=int)
    campaign_id = request.args.get("campaign_id", type=int)

    if not lead_id:
        return jsonify({"error": "missing_lead_id"}), 400

    if not campaign_id:
        return jsonify({"error": "missing_campaign_id"}), 400

    NewsletterES = aliased(Newsletter)
    NewsletterEN = aliased(Newsletter)

    row = (
        db.session.query(
            CampaignRecipient.campaign_id,

            Campaign.name.label("campaign_name"),
            Campaign.campaign_type.label("campaign_type"),
            Campaign.idioma.label("campaign_idioma"),

            Campaign.subject_es.label("subject_es"),
            Campaign.subject_en.label("subject_en"),

            Campaign.sent_at.label("campaign_sent_at"),

            CampaignRecipient.sent_at.label("recipient_sent_at"),
            CampaignRecipient.origen.label("origen"),
            CampaignRecipient.idioma.label("recipient_idioma"),
            CampaignRecipient.send_status.label("send_status"),

            CampaignRecipient.opened_at.label("opened_at"),
            CampaignRecipient.clicked_at.label("clicked_at"),
            CampaignRecipient.click_count.label("click_count"),

            NewsletterES.name.label("newsletter_es_name"),
            NewsletterES.template_s3_path.label("newsletter_es_path"),

            NewsletterEN.name.label("newsletter_en_name"),
            NewsletterEN.template_s3_path.label("newsletter_en_path"),
        )
        .join(
            Campaign,
            Campaign.id == CampaignRecipient.campaign_id
        )
        .outerjoin(
            NewsletterES,
            NewsletterES.id == Campaign.newsletter_es_id
        )
        .outerjoin(
            NewsletterEN,
            NewsletterEN.id == Campaign.newsletter_en_id
        )
        .filter(
            CampaignRecipient.entity_kind == "lead",
            CampaignRecipient.entity_id == lead_id,
            CampaignRecipient.campaign_id == campaign_id
        )
        .first()
    )

    print(
        "campaign_id:", row.campaign_id,
        "subject_es:", row.subject_es,
        "subject_en:", row.subject_en,
        "newsletter_es_name:", row.newsletter_es_name,
        "newsletter_en_name:", row.newsletter_en_name,
        "newsletter_es_path:", row.newsletter_es_path,
        "newsletter_en_path:", row.newsletter_en_path
    )

    if not row:
        return render_template(
            "campanas_detalle_por_lead.html",
            row=None,
            lead_id=lead_id,
            campaign_id=campaign_id,
            s3_bucket=S3_BUCKET
        ), 404

    return render_template(
        "campanas_detalle_por_lead.html",
        row=row,
        lead_id=lead_id,
        campaign_id=campaign_id,
        s3_bucket=S3_BUCKET
    )


@application.route('/campanas_detalle_por_prospecto')
def campanas_detalle_por_prospecto():
    prospectIA_id = request.args.get("prospectIA_id", type=int)
    campaign_id = request.args.get("campaign_id", type=int)

    if not prospectIA_id:
        return jsonify({"error": "missing_prospectIA_id"}), 400

    if not campaign_id:
        return jsonify({"error": "missing_campaign_id"}), 400

    NewsletterES = aliased(Newsletter)
    NewsletterEN = aliased(Newsletter)

    row = (
        db.session.query(
            CampaignRecipient.campaign_id,

            Campaign.name.label("campaign_name"),
            Campaign.campaign_type.label("campaign_type"),
            Campaign.idioma.label("campaign_idioma"),

            Campaign.subject_es.label("subject_es"),
            Campaign.subject_en.label("subject_en"),

            Campaign.sent_at.label("campaign_sent_at"),

            CampaignRecipient.sent_at.label("recipient_sent_at"),
            CampaignRecipient.origen.label("origen"),
            CampaignRecipient.idioma.label("recipient_idioma"),
            CampaignRecipient.send_status.label("send_status"),

            CampaignRecipient.opened_at.label("opened_at"),
            CampaignRecipient.clicked_at.label("clicked_at"),
            CampaignRecipient.click_count.label("click_count"),

            NewsletterES.name.label("newsletter_es_name"),
            NewsletterES.template_s3_path.label("newsletter_es_path"),

            NewsletterEN.name.label("newsletter_en_name"),
            NewsletterEN.template_s3_path.label("newsletter_en_path"),
        )
        .join(
            Campaign,
            Campaign.id == CampaignRecipient.campaign_id
        )
        .outerjoin(
            NewsletterES,
            NewsletterES.id == Campaign.newsletter_es_id
        )
        .outerjoin(
            NewsletterEN,
            NewsletterEN.id == Campaign.newsletter_en_id
        )
        .filter(
            CampaignRecipient.entity_kind == "prospect",
            CampaignRecipient.entity_id == prospectIA_id,
            CampaignRecipient.campaign_id == campaign_id
        )
        .first()
    )

    if not row:
        return render_template(
            "campanas_detalle_por_prospecto.html",
            row=None,
            prospectIA_id=prospectIA_id,
            campaign_id=campaign_id,
            s3_bucket=S3_BUCKET
        ), 404

    print(
        "campaign_id:", row.campaign_id,
        "subject_es:", row.subject_es,
        "subject_en:", row.subject_en,
        "newsletter_es_name:", row.newsletter_es_name,
        "newsletter_en_name:", row.newsletter_en_name,
        "newsletter_es_path:", row.newsletter_es_path,
        "newsletter_en_path:", row.newsletter_en_path
    )

    return render_template(
        "campanas_detalle_por_prospecto.html",
        row=row,
        prospectIA_id=prospectIA_id,
        campaign_id=campaign_id,
        s3_bucket=S3_BUCKET
    )


@application.route("/campana_rebotes")
def campana_rebotes():
    campaign_id = request.args.get("campaign_id", type=int)
    entity_kind = request.args.get("entity_kind", type=str)

    if not campaign_id:
        return jsonify({
            "success": False,
            "error": "missing_campaign_id"
        }), 400

    if entity_kind not in ("lead", "prospect"):
        return jsonify({
            "success": False,
            "error": "invalid_entity_kind"
        }), 400

    campaign = db.session.get(Campaign, campaign_id)

    if not campaign:
        abort(404)

    rows = (
        db.session.query(
            CampaignRecipient.id.label("recipient_id"),
            CampaignRecipient.entity_id,
            CampaignRecipient.entity_kind,

            CampaignRecipient.email,
            CampaignRecipient.pais,
            CampaignRecipient.idioma,
            CampaignRecipient.origen,

            CampaignRecipient.send_status,
            CampaignRecipient.sent_at,
            CampaignRecipient.bounced_at,

            CampaignRecipient.bounce_type,
            CampaignRecipient.bounce_subtype,
            CampaignRecipient.bounce_diagnostic,

            CampaignRecipient.error_message,
            CampaignRecipient.ses_message_id,
        )
        .filter(
            CampaignRecipient.campaign_id == campaign_id,
            CampaignRecipient.entity_kind == entity_kind,
            db.or_(
                CampaignRecipient.send_status == "bounced",
                CampaignRecipient.bounced_at.isnot(None)
            )
        )
        .order_by(
            CampaignRecipient.bounced_at.desc(),
            CampaignRecipient.id.desc()
        )
        .all()
    )

    hard_bounces = sum(
        1 for row in rows
        if (row.bounce_type or "").lower() == "permanent"
    )

    soft_bounces = sum(
        1 for row in rows
        if (row.bounce_type or "").lower() == "transient"
    )

    undetermined_bounces = len(rows) - hard_bounces - soft_bounces

    back_url = url_for(
        "campaign_stats",
        cid=campaign_id,
        entity_kind=entity_kind
    )
    return render_template(
        "campana_rebotes.html",
        campaign=campaign,
        entity_kind=entity_kind,
        rows=rows,
        total_bounces=len(rows),
        hard_bounces=hard_bounces,
        soft_bounces=soft_bounces,
        undetermined_bounces=undetermined_bounces,
        back_url=back_url
    )

@application.route('/contactoFacturaProforma', methods=['GET'])
def contactoFacturaProforma():
    quoteNumber = request.args.get("quoteNumber")
    email = request.args.get("email")
    idioma = request.args.get("idioma")
    pais = request.args.get("pais")
    lead_id = request.args.get("lead_id")
    id_mode= tipo_identificacion_por_pais_texto(pais)

    print ("Renderizando contactoFacturaProforma con:", quoteNumber, email, idioma, pais, id_mode)
    return render_template("contactoFacturaProforma.html",
                           quoteNumber=quoteNumber,
                           email=email,
                           idioma=idioma,
                           pais=pais,
                           id_mode=id_mode,
                           lead_id=lead_id )

@application.route("/actualizar_contacto_y_generar_proforma", methods=["POST"])
def actualizar_contacto():

    from io import BytesIO
    import pycurl
    data = request.get_json(force=True) or {}

    print("ESTOY en actualizar contacto y generar profroma")

    QuoteNo = (data.get("oferta") or "").strip()
    email = (data.get("email") or "").strip()
    idioma = (data.get("idioma") or "").strip()
    pais = (data.get("pais") or "").strip()
    name = (data.get("name") or "").strip()
    Address = (data.get("direccion1") or "").strip()
    Address2 = (data.get("direccion2") or "").strip()
    PostCode = (data.get("codigoPostal") or "").strip()
    City = (data.get("poblacion") or "").strip()
   

    id_mode = data.get("idMode", "").upper()
    ident = (data.get("identificacion") or "").strip()
    n_reg = (data.get("nRegistro") or "").strip()

    VATRegNo = ""
    ForeignRegNo = ""

    if id_mode == "NIF":
        VATRegNo = ident
    elif id_mode == "VAT":
        VATRegNo = ident
    elif id_mode == "REGISTRO":
        ForeignRegNo = ident
    elif id_mode == "VAT+REGISTRO":
        VATRegNo = ident
        ForeignRegNo = n_reg
    else:
        # fallback razonable
        VATRegNo = ident



    print ("Pais recibido en actualizar contacto:", pais)

    

    if not QuoteNo:
        return jsonify({"ok": False, "error": "quoteNumber es obligatorio"}), 400

    

   
            


    print ("URL_ACTUALIZAR_CONTACTO", URL_ACTUALIZAR_CONTACTO)



    api_key = API_KEY
    

    api_url = URL_ACTUALIZAR_CONTACTO

    headers = [
        "x-api-key: " + api_key,
        "Request-Origin: SwaggerBootstrapUi",
        "Accept: application/json",
        "Content-Type: application/json",
    ]

    


    body = json.dumps({
        "QuoteNo": QuoteNo,
        "Name": name,
        "Address": Address,
        "Address2": Address2,
        "PostCode": PostCode,
        "City": City,
        "VATRegNo": VATRegNo,
        "ForeignRegNo": ForeignRegNo,
        "mailorigen": EMAIL_USER,
        "email": email,
        "idioma": idioma,
        
        #"origen": origen,
        "BD": BD,
        "EMAIL_USER": EMAIL_USER,
        "EMAIL_PASSWORD": EMAIL_PASSWORD,
        "URL_CONTACTO": URL_CONTACTO,
        "URL_OFERTAS": URL_OFERTAS,
        "URL_ACTUALIZAR_CONTACTO": URL_ACTUALIZAR_CONTACTO,
        "URL_PROFORMAS": URL_PROFORMAS,
        "URL_FORM_CONTACTO": URL_FORM_CONTACTO,
        "ENVIRONMENT": ENVIRONMENT ,
        "SEND_EMAIL": SEND_EMAIL

    })

    print ("Body a enviar en actualizar contacto:", body)
    

    buffer = BytesIO()
    c = pycurl.Curl()
    c.setopt(c.URL, api_url)
    c.setopt(c.POST, 1)
    c.setopt(c.POSTFIELDS, body)
    c.setopt(pycurl.SSL_VERIFYPEER, 0)
    c.setopt(pycurl.SSL_VERIFYHOST, 0)
    c.setopt(pycurl.CONNECTTIMEOUT, 10)
    c.setopt(pycurl.TIMEOUT, 120)
    c.setopt(c.HTTPHEADER, headers)
    c.setopt(c.WRITEDATA, buffer)

    c.perform()
    status_code = c.getinfo(pycurl.RESPONSE_CODE) or 500
    response_body = buffer.getvalue().decode('utf-8')
    c.close()

    print(f"✅ Respuesta del backend (status {status_code}): {response_body}")

    response = make_response(response_body, status_code)
    response.headers["Content-Type"] = "application/json"
    return response








def _parse_date(s):
    if not s:
        return None
    # Espera 'YYYY-MM-DD' del <input type="date">
    return datetime.strptime(s, "%Y-%m-%d").date()

def _num(x):
    if x is None or x == "":
        return None
    try:
        return Decimal(str(x))
    except (InvalidOperation, ValueError, TypeError):
        return None

def _clip_len(s, n):
    if s is None:
        return None
    return str(s)[:n]


@application.route('/leads', methods=['GET', 'POST'])
def leads():
    if request.method == 'POST':

        # Asegura parseo JSON
        data = request.get_json(force=True) or {}
        
        # Extrae valores del payload
       
        fecha_actual            = data.get('fecha_actual')           # str 'YYYY-MM-DD' o None
        fecha_proyecto          = data.get('fecha_proyecto')
        fecha_proxima_accion    = data.get('fecha_proxima_accion')
        name                    = data.get('name')
        email                   = data.get('email')
        quote_number            = data.get('quote_number')
        idioma                  = data.get('idioma')
        pais                    = data.get('pais')
        incluir_transporte      = data.get('incluir_transporte')    
        importe_transporte      = _num(data.get('importe_transporte'))
        descuento_adicional     = _num(data.get('descuento_adicional'))
        descuento_total         = _num(data.get('descuento_total'))
        cantidad_total          = _num(data.get('cantidad_total'))  
        estado                  = data.get('estado')
        prob_exito_raw          = data.get('probabilidad_exito')
        pistas_perimetrales     = _num(data.get('pistas_perimetrales'))
        pistas_laterales        = _num(data.get('pistas_laterales'))    
        info_tecnica            = _clip_len(data.get('info_tecnica'), 1000)
        info_general            = _clip_len(data.get('info_general'), 1000)
        observaciones           = _clip_len(data.get('observaciones'), 200)

        # Conversión/normalización
        probabilidad_exito = None
        if prob_exito_raw not in (None, ""):
            probabilidad_exito = int(prob_exito_raw)  # lanza ValueError si no es número

        

        # --- Inserción ---
        sql = """
        INSERT INTO lead_forms (
          fecha_actual, fecha_proyecto, fecha_proxima_accion,
          name, email, quote_number, idioma, pais,
          descuento_adicional, descuento_total, cantidad_total,incluir_transporte, importe_transporte,
          probabilidad_exito, pistas_perimetrales, pistas_laterales,
          info_tecnica, info_general, observaciones
        ) VALUES (
          %(fecha_actual)s, %(fecha_proyecto)s, %(fecha_proxima_accion)s,
          %(name)s, %(email)s, %(quote_number)s, %(idioma)s, %(pais)s,
          %(descuento_adicional)s, %(descuento_total)s, %(cantidad_total)s, %(incluir_transporte)s, %(importe_transporte)s,
          %(probabilidad_exito)s, %(pistas_perimetrales)s, %(pistas_laterales)s,
          %(info_tecnica)s, %(info_general)s, %(observaciones)s
        )
        """

        params = {
            #"session_id": session_id,
            "fecha_actual": fecha_actual,
            "fecha_proyecto": fecha_proyecto,
            "fecha_proxima_accion": fecha_proxima_accion,
            "name": name,
            "email": email,
            "quote_number": quote_number,
            "idioma": idioma,
            "pais": pais,
            "descuento_adicional": descuento_adicional,
            "descuento_total": descuento_total,
            "cantidad_total": cantidad_total,
            "probabilidad_exito": probabilidad_exito,
            "incluir_transporte": incluir_transporte,
            "importe_transporte": importe_transporte,   
            "pistas_laterales": pistas_laterales,
            "pistas_perimetrales": pistas_perimetrales,
            "info_tecnica": info_tecnica,
            "info_general": info_general,
            "observaciones": observaciones,
        }

        creds = get_db_credentials("secretoBC/Mysql")

        
        
        dbname = "bc_pruebas" if (BD== "PRUEBAS") else creds["dbname"]

        #print(f"Credenciales obtenidas: {creds}")
        print(f"Conectando a la base de datos con host: {creds['host']}, usuario: {creds['username']}, base de datos: {dbname}")

        conn = pymysql.connect(
            host=creds['host'],
            user=creds['username'],
            password=creds['password'],
            database=dbname,
            port=int(creds.get('port', 3306))
        )

        try:
            
            with conn.cursor() as cur:
                cur.execute(sql, params)
                new_id = cur.lastrowid
            conn.commit()
            return jsonify({"ok": True, "id": new_id, "quote_number": quote_number}), 201

        except pymysql.err.IntegrityError as e:
            conn.rollback()
            # Tip: suele venir (errno, errmsg). Ej. 1048 → NOT NULL
            errno = e.args[0] if e.args else None
            errmsg = e.args[1] if len(e.args) > 1 else str(e)
            print(f"DB IntegrityError {errno}: {errmsg} | params.quote_number={repr(quote_number)}")
            return jsonify({"ok": False, "error": f"MySQL {errno}: {errmsg}"}), 400
        
        except pymysql.err.Error as e:
            conn.rollback()
            print(f"DB Error: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

        finally:
            try:
                conn.close()
            except Exception:
                pass

        return make_response(jsonify({"ok": True, "id": new_id}), 201)

    # GET opcional: podrías listar últimos leads
    #return make_response(jsonify({"ok": True, "msg": "Use POST para crear leads"}), 200)
        
       
    # Si es GET, renderiza el formulario

    incluir_transporte = request.args.get('incluir_transporte')        # 'true' / 'false' / None
    importe_transporte = request.args.get('importe_transporte')        # '400' / None
    name                = request.args.get('name')
    email               = request.args.get('email')
    idioma              = request.args.get('idioma')
    pais                = request.args.get('pais')
    descuento_adicional = request.args.get('descuento_adicional')
    descuento_total     = request.args.get('descuento_total')
    cantidad_total      = request.args.get('cantidad_total')
    quote_number        = request.args.get('quoteNumber')  # ojo: en la URL es quoteNumber
    pistas_perimetrales = request.args.get('pistas_perimetrales')
    pistas_laterales    = request.args.get('pistas_laterales')

    # si quieres, convierte incluir_transporte a boolean:
    if incluir_transporte is not None:
        incluir_transporte = incluir_transporte.lower() == 'true'
    return render_template(
            'leads.html',
            incluir_transporte=incluir_transporte,
            importe_transporte=importe_transporte,
            name=name,
            email=email,
            idioma=idioma,
            pais=pais,
            descuento_adicional=descuento_adicional,
            descuento_total=descuento_total,
            cantidad_total=cantidad_total,
            quote_number=quote_number,
            pistas_perimetrales=pistas_perimetrales,
            pistas_laterales=pistas_laterales,
        )

@application.route("/consultar_leads", methods=["GET", "POST"])
def consultar_leads():

    estado = request.args.get("estado")
    modo = request.args.get("modo", "leads")
    campaign_id = request.args.get("campaign_id", type=int)
    campaign_name = request.args.get("campaign_name", "")

    # Evita aceptar modos no previstos
    modos_permitidos = {
        "leads",
        "aperturas",
        "clicks",
        "rebotes",
        "bajas",
    }

    if modo not in modos_permitidos:
        modo = "leads"

    titulos = {
        "leads": "Consultar leads",
        "aperturas": "Leads que han abierto la campaña: ",
        "clicks": "Leads que han hecho clic en la campaña: ",
        "rebotes": "Leads con rebote en la campaña:",
        "bajas": "Leads dados de baja en la campaña:",
    }

    titulo = titulos[modo]

    creds = get_db_credentials("secretoBC/Mysql")
    dbname = "bc_pruebas" if BD == "PRUEBAS" else creds["dbname"]

    print(
        f"Conectando a la base de datos con host: {creds['host']}, "
        f"usuario: {creds['username']}, base de datos: {dbname}"
    )

    conn = pymysql.connect(
        host=creds["host"],
        user=creds["username"],
        password=creds["password"],
        database=dbname,
        port=int(creds.get("port", 3306)),
    )

    from pymysql.cursors import DictCursor

    # Condiciones para lead_forms
    condiciones_lead = []
    params_lead = []

    if estado and estado != "Todos":
        condiciones_lead.append("lf.estado = %s")
        params_lead.append(estado)

    # En el informe de bajas se muestran solo leads dados de baja
    if modo == "bajas":
        condiciones_lead.append("lf.unsubscribed = 1")

    where_lead_sql = ""

    if condiciones_lead:
        where_lead_sql = (
            "WHERE " + " AND ".join(condiciones_lead)
        )

    # Condiciones dentro de campaign_recipients
    condiciones_campaign = [
        "entity_kind = 'lead'"
    ]
    params_campaign = []

    if campaign_id:
        condiciones_campaign.append("campaign_id = %s")
        params_campaign.append(campaign_id)

    if modo == "aperturas":
        condiciones_campaign.append("opened_at IS NOT NULL")

    elif modo == "clicks":
        condiciones_campaign.append("clicked_at IS NOT NULL")

    elif modo == "rebotes":
        condiciones_campaign.append("bounced_at IS NOT NULL")

    where_campaign_sql = (
        "WHERE " + " AND ".join(condiciones_campaign)
    )

    # En informes de actividad solo queremos leads que tengan coincidencia
    if modo in ("aperturas", "clicks", "rebotes", "bajas"):
        tipo_join = "INNER JOIN"
    else:
        tipo_join = "LEFT JOIN"

    # Los parámetros aparecen primero en la subconsulta y luego en el WHERE exterior
    params = params_campaign + params_lead

    rows = []

    try:
        with conn.cursor(DictCursor) as cur:

            sql = f"""
                SELECT
                    lf.id,
                    lf.fecha_actual,
                    lf.fecha_proyecto,
                    lf.fecha_proxima_accion,
                    lf.name,
                    lf.pais,
                    lf.tipo_lead,
                    lf.quote_number,
                    lf.cantidad_total,
                    lf.descuento_total,

                    COALESCE(lf.pistas_laterales, 0)
                        + COALESCE(lf.pistas_perimetrales, 0)
                        AS pistas_total,

                    lf.probabilidad_exito,
                    lf.incluir_transporte,
                    lf.importe_transporte,
                    lf.estado,

                    lf.renting_solicitado,
                    lf.fecha_solicitud_renting,
                    lf.proforma_solicitada,
                    lf.fecha_solicitud_proforma,

                    lf.unsubscribed,
                    lf.email_suppressed,
                    lf.email_suppressed_reason,

                    COUNT(cr.campaign_id) AS total_campanas,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN cr.opened_at IS NOT NULL THEN 1
                                ELSE 0
                            END
                        ),
                        0
                    ) AS total_abiertas,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN cr.clicked_at IS NOT NULL THEN 1
                                ELSE 0
                            END
                        ),
                        0
                    ) AS total_campanas_clicadas,

                    COALESCE(
                        SUM(cr.click_count),
                        0
                    ) AS total_clicks,

                    MAX(cr.opened_at) AS ultima_apertura,
                    MAX(cr.clicked_at) AS ultimo_click,
                    MAX(cr.bounced_at) AS ultimo_rebote,

                    MAX(cr.bounce_type) AS bounce_type,
                    MAX(cr.bounce_subtype) AS bounce_subtype,
                    MAX(cr.bounce_diagnostic) AS bounce_diagnostic

                FROM lead_forms lf

                {tipo_join} (
                    SELECT
                        entity_id,
                        campaign_id,

                        MAX(opened_at) AS opened_at,
                        MAX(clicked_at) AS clicked_at,
                        MAX(bounced_at) AS bounced_at,

                        MAX(bounce_type) AS bounce_type,
                        MAX(bounce_subtype) AS bounce_subtype,
                        MAX(bounce_diagnostic) AS bounce_diagnostic,

                        MAX(
                            COALESCE(click_count, 0)
                        ) AS click_count

                    FROM campaign_recipients

                    {where_campaign_sql}

                    GROUP BY
                        entity_id,
                        campaign_id
                ) cr
                    ON cr.entity_id = lf.id

                {where_lead_sql}

                GROUP BY
                    lf.id,
                    lf.fecha_actual,
                    lf.fecha_proyecto,
                    lf.fecha_proxima_accion,
                    lf.name,
                    lf.pais,
                    lf.tipo_lead,
                    lf.quote_number,
                    lf.cantidad_total,
                    lf.descuento_total,
                    lf.pistas_laterales,
                    lf.pistas_perimetrales,
                    lf.probabilidad_exito,
                    lf.incluir_transporte,
                    lf.importe_transporte,
                    lf.estado,
                    lf.renting_solicitado,
                    lf.fecha_solicitud_renting,
                    lf.proforma_solicitada,
                    lf.fecha_solicitud_proforma,
                    lf.unsubscribed,
                    lf.email_suppressed,
                    lf.email_suppressed_reason

                ORDER BY
                    lf.fecha_actual DESC,
                    lf.id DESC
            """

            cur.execute(sql, params)
            rows = cur.fetchall()

    finally:
        conn.close()

    print(f"Modo del informe: {modo}")
    print(f"Campaña: {campaign_id}")
    print(f"Leads obtenidos: {len(rows)}")

    return render_template(
        "consultar_leads.html",
        leads=rows,
        estado=estado,
        modo=modo,
        titulo=titulo,
        campaign_id=campaign_id,
        campaign_name=campaign_name,
    )


def db_get_lead(lead_id):
    creds = get_db_credentials("secretoBC/Mysql")
    

    dbname = "bc_pruebas" if (BD== "PRUEBAS") else creds["dbname"]
    #print(f"Credenciales obtenidas: {creds}")
    print(f"Conectando a la base de datos con host: {creds['host']}, usuario: {creds['username']}, base de datos: {dbname}")

    conn = pymysql.connect(
        host=creds['host'],
        user=creds['username'],
        password=creds['password'],
        database=dbname,
        port=int(creds.get('port', 3306))
    )


    from pymysql.cursors import DictCursor

    where_sql = "WHERE id = %s"
    params = []
   
    params.append(lead_id)

   
    try:
        with conn.cursor(DictCursor) as cur:
            sql = f"""
                SELECT
                id,
                fecha_actual,
                fecha_proyecto,
                fecha_proxima_accion,
                name,
                email,
                idioma,
                pais,
                tipo_lead,
                quote_number,
                cantidad_total,
                descuento_adicional,
                descuento_total,
                pistas_perimetrales,
                pistas_laterales,
                probabilidad_exito,
                incluir_transporte,
                importe_transporte,
                info_tecnica,
                info_general,
                observaciones,
                estado
                FROM lead_forms
                {where_sql}
                ORDER BY fecha_actual DESC, id DESC
            """
            print ("sql",sql)
            print ("params", params)
            cur.execute(sql, params)   # 👈 pasa params (aunque esté vacío)
            rows = cur.fetchall()
    finally:
        conn.close()

    
    print(f"Leads obtenidos: {len(rows)}")
    print("Leads:", rows)
                               
    return (rows[0] if rows else None)  # Devuelve el primer lead o None si no existe

def db_get_prospect(prospectIA_id):
    creds = get_db_credentials("secretoBC/Mysql")

    dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

    print(
        f"Conectando a la base de datos con host: {creds['host']}, "
        f"usuario: {creds['username']}, base de datos: {dbname}"
    )

    conn = pymysql.connect(
        host=creds['host'],
        user=creds['username'],
        password=creds['password'],
        database=dbname,
        port=int(creds.get('port', 3306))
    )

    from pymysql.cursors import DictCursor

    where_sql = "WHERE id = %s"
    params = [prospectIA_id]

    try:
        with conn.cursor(DictCursor) as cur:
            sql = f"""
                SELECT
                    id,
                    fecha,
                    idioma,
                    pais,
                    email,
                    club,
                    estado,
                    propietario,
                    num_pistas,
                    tipo,
                    web,
                    youtube,
                    instagram,
                    linkedin_club,
                    linkedin_propietario,
                    booking_app,
                    proveedor_pistas,
                    unsubscribed,
                    unsubscribed_at
                FROM prospects_IA
                {where_sql}
            """
            print("sql", sql)
            print("params", params)

            cur.execute(sql, params)
            rows = cur.fetchall()

    finally:
        conn.close()

    print(f"Prospectos obtenidos: {len(rows)}")
    print("Prospectos:", rows)

    return (rows[0] if rows else None)


def db_update_lead(lead):
    creds = get_db_credentials("secretoBC/Mysql")
   

    dbname = "bc_pruebas" if (BD== "PRUEBAS") else creds["dbname"]

    #print(f"Credenciales obtenidas: {creds}")
    print(f"Conectando a la base de datos con host: {creds['host']}, usuario: {creds['username']}, base de datos: {dbname}")

    conn = pymysql.connect(
        host=creds['host'],
        user=creds['username'],
        password=creds['password'],
        database=dbname,
        port=int(creds.get('port', 3306))
    )


    from pymysql.cursors import DictCursor

    try:
        with conn.cursor(DictCursor) as cur:
            sql = """
                UPDATE lead_forms
                SET
                  fecha_actual          = %s,
                  fecha_proyecto        = %s,
                  fecha_proxima_accion  = %s,
                  probabilidad_exito    = %s,
                  tipo_lead             = %s,
                  info_tecnica          = %s,
                  info_general          = %s,
                  observaciones         = %s,
                  estado                = %s
                WHERE id = %s
            """
            params = [
                lead.fecha_actual,
                lead.fecha_proyecto,
                lead.fecha_proxima_accion,
                lead.probabilidad_exito,
                lead.tipo_lead,
                lead.info_tecnica,
                lead.info_general,
                lead.observaciones,
                lead.estado,
                lead.id   # 👈 muy importante que el id vaya al final
            ]
            print("SQL:", sql)
            print("Params:", params)

            cur.execute(sql, params)
            conn.commit()   # 👈 confirma cambios
            rows = cur.rowcount
            print("Filas actualizadas:", rows)
    finally:
        conn.close()

    
   



@application.route('/lead_manage', methods=['GET', 'POST'])
def lead_manage():

    

    try:
        if request.method == "POST":
            from types import SimpleNamespace
            
            try:
                # Asegura parseo JSON
                data = request.get_json(force=True) or {}
                
                # Extrae valores del payload
                id_                     = data.get('id')
                fecha_actual            = data.get('fecha_actual')           # str 'YYYY-MM-DD' o None
                fecha_proyecto          = data.get('fecha_proyecto')
                fecha_proxima_accion    = data.get('fecha_proxima_accion')
                estado                  = data.get('estado')
                tipo_lead               = data.get('tipo_lead')
                prob_exito_raw          = data.get('probabilidad_exito')
                incluir_transporte      = data.get('incluir_transporte')
                importe_transporte      = _num(data.get('importe_transporte'))
                info_tecnica            = _clip_len(data.get('info_tecnica'), 1000)
                info_general            = _clip_len(data.get('info_general'), 1000)
                observaciones           = _clip_len(data.get('observaciones'), 200)

                # Conversión/normalización
                probabilidad_exito = None
                if prob_exito_raw not in (None, ""):
                    probabilidad_exito = int(prob_exito_raw)  # lanza ValueError si no es número

                # Valida mínimos
                if not id_:
                    return jsonify({"ok": False, "message": "Falta id"}), 400

                # Crea 'lead' como un objeto con atributos 
                lead = SimpleNamespace(
                    id=id_,
                    fecha_actual=fecha_actual,
                    fecha_proyecto=fecha_proyecto,
                    fecha_proxima_accion=fecha_proxima_accion,
                    estado=estado,
                    tipo_lead=tipo_lead,
                    probabilidad_exito=probabilidad_exito,
                    incluir_transporte=incluir_transporte,
                    importe_transporte=importe_transporte,
                    info_tecnica=info_tecnica,
                    info_general=info_general,
                    observaciones=observaciones
                )

                # Logs correctos (usa las variables definidas)
                print("📥 Datos recibidos:")
                print(f"id: {lead.id}")
                print(f"Fecha actual: {lead.fecha_actual}")
                print(f"Fecha proyecto: {lead.fecha_proyecto}")
                print(f"Fecha próxima acción: {lead.fecha_proxima_accion}")
                print(f"Estado: {lead.estado}")
                print(f"Tipo de lead: {lead.tipo_lead}")
                print(f"Probabilidad de éxito: {lead.probabilidad_exito}")
                print(f"Información técnica: {lead.info_tecnica}")
                print(f"Información general: {lead.info_general}")
                print(f"Observaciones: {lead.observaciones}")

                # Persistencia
                db_update_lead(lead)

                return jsonify({"ok": True})

            except Exception as e:
                application.logger.exception("Error en lead_manage")
                return jsonify({"ok": False, "message": str(e)}), 400

        # GET
        else:
            lead_id = request.args.get("lead_id")

            modo = request.args.get("modo", "leads")
            campaign_id = request.args.get("campaign_id", type=int)
            estado_origen = request.args.get("estado", "Sin calificar")
            campaign_name = request.args.get("campaign_name")

            if not lead_id:
                return jsonify({"error": "missing_lead_id"}), 400

            lead = db_get_lead(lead_id)

            if not lead:
                return render_template(
                    "lead_not_found.html",
                    lead_id=lead_id
                ), 404

            return render_template(
                "lead_manage.html",
                lead=lead,
                modo=modo,
                campaign_id=campaign_id,
                estado_origen=estado_origen,
                campaign_name=campaign_name
            )

    except Exception as e:
        application.logger.exception("Error inesperado en lead_manage")
        # Devolver algo incluso en error
        return jsonify({"error": "internal", "detail": str(e)}), 500
    


@application.route('/prospectsIA_manage', methods=['GET', 'POST'])
def prospectsIA_manage():
    try:

        if request.method == "POST":
            try:
                data = request.get_json(force=True) or {}

                id_ = data.get("id")
                fecha = data.get("fecha")
                estado = data.get("estado")
                idioma = data.get("idioma")
                pais = data.get("pais")
                num_pistas = _num(data.get("num_pistas"))
                tipo = data.get("tipo")
                youtube = _clip_len(data.get("youtube"), 100)
                instagram = _clip_len(data.get("instagram"), 100)
                propietario = _clip_len(data.get("propietario"), 100)
                web = _clip_len(data.get("web"), 255)
                linkedin_club = _clip_len(data.get("linkedin_club"), 100)
                linkedin_propietario = _clip_len(data.get("linkedin_propietario"), 100)
                booking_app = _clip_len(data.get("booking_app"), 100)
                proveedor_pistas = _clip_len(data.get("proveedor_pistas"), 100)

                if not id_:
                    return jsonify({"ok": False, "message": "Falta id"}), 400

                if not fecha:
                    return jsonify({"ok": False, "message": "Falta fecha"}), 400

                if not estado:
                    return jsonify({"ok": False, "message": "Falta estado"}), 400

                if not idioma:
                    return jsonify({"ok": False, "message": "Falta idioma"}), 400

                if not pais:
                    return jsonify({"ok": False, "message": "Falta país"}), 400

                estados_validos = {"operativo", "renovacion", "concepto", "proyecto"}
                if estado not in estados_validos:
                    return jsonify({"ok": False, "message": "Estado no válido"}), 400

                prospectIA = SimpleNamespace(
                    id=id_,
                    fecha=fecha,
                    estado=estado,
                    idioma=idioma,
                    pais=pais,
                    num_pistas=num_pistas,
                    tipo=tipo,
                    youtube=youtube,
                    instagram=instagram,
                    propietario=propietario,
                    web=web,
                    linkedin_club=linkedin_club,
                    linkedin_propietario=linkedin_propietario,
                    booking_app=booking_app,
                    proveedor_pistas=proveedor_pistas,
                )

                print("📥 Datos recibidos prospectIA:")
                print(f"id: {prospectIA.id}")
                print(f"fecha: {prospectIA.fecha}")
                print(f"estado: {prospectIA.estado}")
                print(f"idioma: {prospectIA.idioma}")
                print(f"pais: {prospectIA.pais}")
                print(f"num_pistas: {prospectIA.num_pistas}")
                print(f"tipo: {prospectIA.tipo}")
                print(f"youtube: {prospectIA.youtube}")
                print(f"instagram: {prospectIA.instagram}")
                print(f"propietario: {prospectIA.propietario}")
                print(f"web: {prospectIA.web}")
                print(f"linkedin_club: {prospectIA.linkedin_club}")
                print(f"linkedin_propietario: {prospectIA.linkedin_propietario}")
                print(f"booking_app: {prospectIA.booking_app}")
                print(f"proveedor_pistas: {prospectIA.proveedor_pistas}")

                db_update_prospectIA(prospectIA)

                return jsonify({"ok": True})

            except Exception as e:
                print("❌ Error en POST prospectsIA_manage:", e)
                return jsonify({"ok": False, "message": str(e)}), 400

        prospectIA_id = request.args.get("prospectIA_id")

        modo = request.args.get("modo", "leads")
        campaign_id = request.args.get("campaign_id", type=int)
        estado_origen = request.args.get("estado", "Sin calificar")
        campaign_name = request.args.get("campaign_name")

        if not prospectIA_id:
            flash("Falta prospectIA_id", "error")
            return redirect(url_for("consultar_prospectos_IA"))

        prospectIA = db_get_prospect(prospectIA_id)

        if not prospectIA:
            flash("Prospecto no encontrado", "error")
            return redirect(url_for("consultar_prospectos_IA"))

        if prospectIA.get("fecha"):
            prospectIA["fecha"] = prospectIA["fecha"].isoformat()

        if prospectIA.get("unsubscribed_at"):
            prospectIA["unsubscribed_at"] = prospectIA["unsubscribed_at"].isoformat()

        print(f"ProspectIA obtenido para id {prospectIA_id}: {prospectIA}")


        return render_template(
            "prospectsIA_manage.html",
            prospectIA=prospectIA,
            modo=modo,
            campaign_id=campaign_id,
            estado_origen=estado_origen,
            campaign_name=campaign_name
        )

        # return render_template("prospectsIA_manage.html", prospectIA=prospectIA)

    except Exception as e:
        print("Error inesperado en prospectIA_manage")
        import traceback
        traceback.print_exc()
        flash(f"Error inesperado: {e}", "error")
        return redirect(url_for("consultar_prospectos_IA"))

def db_update_prospectIA(prospectIA):
    creds = get_db_credentials("secretoBC/Mysql")
    dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

    print(
        f"Conectando a la base de datos con host: {creds['host']}, "
        f"usuario: {creds['username']}, base de datos: {dbname}"
    )

    conn = pymysql.connect(
        host=creds["host"],
        user=creds["username"],
        password=creds["password"],
        database=dbname,
        port=int(creds.get("port", 3306)),
        autocommit=False
    )

    try:
        with conn.cursor() as cur:
            sql = """
                UPDATE prospects_IA
                SET
                    fecha = %s,
                    estado = %s,
                    idioma = %s,
                    pais = %s,
                    num_pistas = %s,
                    tipo = %s,
                    youtube = %s,
                    instagram = %s,
                    propietario = %s,
                    web = %s,
                    linkedin_club = %s,
                    linkedin_propietario = %s,
                    booking_app = %s,
                    proveedor_pistas = %s
                WHERE id = %s
            """

            params = (
                prospectIA.fecha,
                prospectIA.estado,
                prospectIA.idioma,
                prospectIA.pais,
                prospectIA.num_pistas,
                prospectIA.tipo,
                prospectIA.youtube,
                prospectIA.instagram,
                prospectIA.propietario,
                prospectIA.web,
                prospectIA.linkedin_club,
                prospectIA.linkedin_propietario,
                prospectIA.booking_app,
                prospectIA.proveedor_pistas,
                prospectIA.id,
            )

            print("sql UPDATE prospects_IA:", sql)
            print("params:", params)

            cur.execute(sql, params)
            conn.commit()

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()

@application.route('/prospectos_IA', methods=['GET', 'POST'])
def prospectos_IA():
    session.clear()  # Elimina todos los datos de sesión
    return redirect(url_for('login'))  # Cambiá 'login' por tu vista de inicio o login

#@application.route('/campanas', methods=['GET', 'POST'])
#def campanas():
#    session.clear()  # Elimina todos los datos de sesión
#    return redirect(url_for('login'))  # Cambiá 'login' por tu vista de inicio o login

from datetime import datetime,timezone

def save_newsletter_db(name, idioma,template_s3_path):

    creds = get_db_credentials("secretoBC/Mysql")
    dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

    conn = pymysql.connect(
        host=creds["host"],
        user=creds["username"],
        password=creds["password"],
        database=dbname,
        port=int(creds.get("port", 3306)),
        cursorclass=pymysql.cursors.DictCursor
    )
    try:
   
   
        cursor = conn.cursor()
  

        now = datetime.now(timezone.utc)

        cursor.execute(
            "SELECT id FROM newsletters WHERE name = %s AND lang=%s",
            (name,idioma)
        )
        row = cursor.fetchone()

        if row:
            cursor.execute("""
                UPDATE newsletters
                SET                
                    template_s3_path = %s,
                
                    lang = %s,
                    updated_at = %s
                WHERE name = %s
            """, (
                
                template_s3_path,
                idioma,
                now,
                name
            ))
        else:
            cursor.execute("""
                INSERT INTO newsletters
                (name,template_s3_path, lang,created_at, updated_at)
                VALUES (%s, %s, %s, %s,%s)
            """, (
                name,
                template_s3_path,
                idioma,
                now,
                now
            ))

        conn.commit()
    finally:
        cursor.close()
        conn.close()

@application.route('/upload_template_email', methods=['GET', 'POST'])
def upload_template_email():
    if request.method == "POST":
        try:
            payload = request.get_json(force=True, silent=False)
        except Exception as e:
            print("[ERROR] JSON parse:", repr(e))
            return jsonify({"error": "JSON inválido"}), 400

        print("[DEBUG] payload keys:", list(payload.keys()))

        name = (payload.get("name") or "").strip()
        eml_b64 = (payload.get("eml_base64") or "").strip()
        zip_b64 = (payload.get("zip_base64") or "").strip()
        html_in = (payload.get("html") or "").strip()
        lang = (payload.get("lang") or "es").strip().lower()
        template_type = (payload.get("template_type") or "email").strip().lower()

        if template_type not in ("email", "newsletter"):
            return jsonify({"error": "template_type debe ser 'email' o 'newsletter'"}), 400

        if not name:
            return jsonify({"error": "Campo 'name' es obligatorio"}), 400

        if not eml_b64 and not html_in and not zip_b64:
            return jsonify({"error": "Envía 'eml_base64', 'zip_base64' o 'html'"}), 400

        slug = slugify(name)

        # salida local separada por tipo
        out_dir = os.path.join("output", template_type, slug, lang)

        # prefijo S3 separado por tipo
        base_prefix = "emails/templates" if template_type == "email" else "newsletters/templates"

        def ensure_html(s: str) -> str:
            if "<" in s and ">" in s:
                return s
            blocks = [f"<p>{line.strip()}</p>" for line in s.split("\n\n") if line.strip()]
            return "<html><body>" + "\n".join(blocks) + "</body></html>"

        attachments = []
        images = []

        # =========================================================
        # 1) ENTRADA DESDE .EML
        # =========================================================
        if eml_b64:
            print("[DEBUG] usando flujo EML; eml_base64 len:", len(eml_b64))
            try:
                eml_bytes = base64.b64decode(eml_b64, validate=True)
                print("[DEBUG] eml_bytes len:", len(eml_bytes))
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({"ok": False, "where": "base64", "error": str(e)}), 400

            try:
                extracted = extract_html_inline_and_attachments_from_eml_bytes(
                    eml_bytes,
                    slug,
                    lang,
                    append_unreferenced_images=True,
                    base_prefix=base_prefix   # <- importante
                )
            except TypeError:
                # compatibilidad si tu helper aún no acepta base_prefix
                extracted = extract_html_inline_and_attachments_from_eml_bytes(
                    eml_bytes,
                    slug,
                    lang,
                    append_unreferenced_images=True
                )
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({"ok": False, "where": "extractor", "error": str(e)}), 400

            if not extracted or not isinstance(extracted, dict):
                return jsonify({
                    "ok": False,
                    "where": "extractor",
                    "error": "Extractor devolvió None o tipo no dict"
                }), 400

            if not extracted.get("html"):
                return jsonify({
                    "ok": False,
                    "where": "extractor",
                    "error": "Extractor sin parte HTML"
                }), 400

            html_final = unescape_pre_wrapped_html(extracted["html"])
            

        # =====================================================
        # NEWSLETTER + ZIP -> html + assets
        # =====================================================
        if template_type == "newsletter" and zip_b64:
            try:
                zip_bytes = base64.b64decode(zip_b64, validate=True)
            except Exception as e:
                return jsonify({"error": f"ZIP base64 inválido: {e}"}), 400

            try:
                html_final, uploaded_assets = _upload_zip_assets_and_rewrite_html(
                    zip_bytes=zip_bytes,
                    slug=slug,
                    lang=lang,
                    base_prefix=base_prefix
                )
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({"error": f"No se pudo procesar el ZIP: {e}"}), 400

            out_dir = os.path.join("output", template_type, slug, lang)
            os.makedirs(out_dir, exist_ok=True)

            original_key = f"{base_prefix}/{slug}/{lang}/original.html"
            template_key = f"{base_prefix}/{slug}/{lang}/template.html"
            schema_key = f"{base_prefix}/{slug}/{lang}/schema.json"
            manifest_key = f"{base_prefix}/{slug}/manifest.json"

            put_public_s3(
                original_key,
                html_final.encode("utf-8"),
                "text/html; charset=utf-8",
                cache_seconds=0
            )

            put_public_s3(
                template_key,
                html_final.encode("utf-8"),
                "text/html; charset=utf-8",
                cache_seconds=0
            )

            put_public_s3(
                schema_key,
                json.dumps({}, ensure_ascii=False).encode("utf-8"),
                "application/json",
                cache_seconds=0
            )

            existing_manifest = {}

            try:
                existing_txt = s3_get_text(manifest_key)
                if existing_txt:
                    existing_manifest = json.loads(existing_txt)
            except Exception:
                existing_manifest = {}

            manifest = existing_manifest or {}

            manifest["slug"] = slug
            manifest["display_name"] = name
            manifest["type"] = "newsletter"
            manifest["base_prefix"] = base_prefix

            shared = manifest.setdefault("shared", {})
            existing_assets = shared.get("assets") or []

            assets_by_key = {
                a.get("key"): a for a in existing_assets if a.get("key")
            }
            for a in uploaded_assets:
                if a.get("key"):
                    assets_by_key[a["key"]] = a

            shared["assets"] = list(assets_by_key.values())

            languages = manifest.setdefault("languages", {})
            languages[lang] = {
                "paths": {
                    "original": original_key,
                    "html": template_key,
                    "schema": schema_key
                }
            }

            put_public_s3(
                manifest_key,
                json.dumps(manifest, indent=2, ensure_ascii=False).encode("utf-8"),
                "application/json",
                cache_seconds=0
            )

            put_public_s3(
                manifest_key,
                json.dumps(manifest, indent=2, ensure_ascii=False).encode("utf-8"),
                "application/json",
                cache_seconds=0
            )
            save_newsletter_db(
                name=name,
                idioma= lang,
                template_s3_path=template_key
            )

            return jsonify({
                "template_id": str(uuid.uuid4()),
                "name": name,
                "slug": slug,
                "lang": lang,
                "type": "newsletter",
                "assets_uploaded": len(uploaded_assets),
                "paths": {
                    "original": original_key,
                    "html": template_key,
                    "schema": schema_key,
                    "manifest": manifest_key
                }
            }), 200
            attachments = extracted.get("attachments", [])
            images = extracted.get("images", [])

            print("[DEBUG] extractor.debug:", extracted.get("debug", {}))

            put_public_s3(
                f"{base_prefix}/{slug}/{lang}/original.html",
                html_final.encode("utf-8"),
                "text/html; charset=utf-8",
                cache_seconds=0
            )

        # =========================================================
        # 2) ENTRADA DESDE HTML
        # =========================================================
        else:
            print("[DEBUG] usando flujo HTML")
            html_final = ensure_html(html_in)

            payload_attachments = payload.get("attachments") or []

            # Resolver cid: si el front manda binarios
            if payload_attachments:
                try:
                    html_final, resolved_info = resolve_cid_with_attachments(
                        html_final,
                        slug,
                        payload_attachments,
                        base_prefix=base_prefix
                    )
                except TypeError:
                    html_final, resolved_info = resolve_cid_with_attachments(
                        html_final,
                        slug,
                        payload_attachments
                    )

            # Rehost de imágenes a S3
            try:
                html_final, stats = rehost_images_under_template_from_html(
                    html_final,
                    slug,
                    base_prefix=base_prefix
                )
            except TypeError:
                html_final, stats = rehost_images_under_template_from_html(
                    html_final,
                    slug
                )

            print("[DEBUG] rehost stats:", stats)
            images = stats.get("uploaded", [])

            if any(
                s.get("reason") == "cid_in_html_send_eml"
                for s in (stats.get("skipped") or [])
            ) and not payload_attachments:
                return jsonify({
                    "error": "El HTML contiene imágenes cid:. Sube el .eml (eml_base64) o envía attachments[] con los binarios."
                }), 400

            # Subir adjuntos extra
            if payload_attachments:
                try:
                    html_final, added_files = insert_extra_files_into_html(
                        html_final,
                        slug,
                        payload_attachments,
                        base_prefix=base_prefix
                    )
                except TypeError:
                    html_final, added_files = insert_extra_files_into_html(
                        html_final,
                        slug,
                        payload_attachments
                    )
                attachments = added_files

        # =========================================================
        # 3) SEPARAR IMÁGENES DE FIRMA
        # =========================================================
        SIGNATURE_MAX_BYTES = 30 * 1024

        signature_images = []
        content_images = []

        for img in images or []:
            size = img.get("size") or img.get("filesize") or img.get("length") or 0
            try:
                size = int(size)
            except (TypeError, ValueError):
                size = 0

            if size and size < SIGNATURE_MAX_BYTES:
                signature_images.append(img)
            else:
                content_images.append(img)

        images = content_images

        # =========================================================
        # 4) SEPARAR CUERPO Y FIRMA
        # =========================================================
        body_html, signature_html = split_body_and_signature(html_final)
        signature_html = clean_signature_images(signature_html)
        html_final = body_html

        # =========================================================
        # 5) CONSTRUIR FRAMEWORK INTERNO
        # =========================================================
        try:
            result = build_framework(
                input_path_or_html=html_final,
                out_dir=out_dir,
                slug=slug,
                lang=lang,
                upload_to_s3=True,
                display_name=name,
                lang_attachments=attachments,
                base_prefix=base_prefix,      # <- importante
                template_type=template_type   # <- importante
            )
        except TypeError:
            # compatibilidad si build_framework aún no acepta esos parámetros
            result = build_framework(
                input_path_or_html=html_final,
                out_dir=out_dir,
                slug=slug,
                lang=lang,
                upload_to_s3=True,
                display_name=name,
                lang_attachments=attachments
            )

        # =========================================================
        # 6) GUARDAR FIRMA COMO PARTIAL
        # =========================================================
        # Si quieres firma solo para emails, deja esta condición:
        # if template_type == "email" and signature_html:
        if signature_html:
            signature_key = f"{base_prefix}/{slug}/partials/signature.html"

            os.makedirs(os.path.join("output", template_type, slug, "partials"), exist_ok=True)

            local_signature_path = os.path.join(
                "output", template_type, slug, "partials", "signature.html"
            )

            with open(local_signature_path, "w", encoding="utf-8") as f:
                f.write(signature_html)

            put_public_s3(
                signature_key,
                signature_html.encode("utf-8"),
                "text/html; charset=utf-8",
                cache_seconds=0
            )

        # =========================================================
        # 7) ACTUALIZAR MANIFEST
        # =========================================================
        manifest_path = result.get("manifest")
        if manifest_path and os.path.exists(manifest_path):
            with open(manifest_path, "r", encoding="utf-8") as f:
                m = json.load(f)
        else:
            m = {}

        m["type"] = template_type
        m["name"] = name
        m["slug"] = slug
        m["lang"] = lang
        m["attachments"] = attachments
        m["images_uploaded"] = images

        if signature_html:
            shared = m.setdefault("shared", {})
            partials = shared.setdefault("partials", {})
            partials["signature_html"] = f"{base_prefix}/{slug}/partials/signature.html"

        os.makedirs(out_dir, exist_ok=True)
        with open(os.path.join(out_dir, "manifest.json"), "w", encoding="utf-8") as f:
            json.dump(m, f, indent=2, ensure_ascii=False)

        return jsonify({
            "template_id": str(uuid.uuid4()),
            "name": name,
            "slug": slug,
            "lang": lang,
            "type": template_type,
            "base_prefix": base_prefix,
            "paths": result,
            "attachments": attachments,
            "images": images,
            "signature_html": signature_html
        }), 200

    return render_template('upload_template_email.html')

@application.route('/upload_template_email1', methods=['GET', 'POST'])
def upload_template_email1():
    if request.method == "POST":


        try:
            payload = request.get_json(force=True, silent=False)
        except Exception as e:
            print("[ERROR] JSON parse:", repr(e))
            return jsonify({"error": "JSON inválido"}), 400

        print("[DEBUG] payload keys:", list(payload.keys()))
        name = (payload.get("name") or "").strip()
        eml_b64 = (payload.get("eml_base64") or "").strip()
        html_in = (payload.get("html") or "").strip()
        if not name:
            return jsonify({"error": "Campo 'name' es obligatorio"}), 400
        if not eml_b64 and not html_in:
            return jsonify({"error": "Envía 'eml_base64' (archivo .eml) o 'html'"}), 400
        lang = (payload.get("lang") or "es").strip().lower()
        print("[DEBUG] lang payload:", lang)

        
        lang = (payload.get("lang") or "es").strip().lower()
        print("[DEBUG] lang payload:", lang)

        slug = slugify(name)
        out_dir = os.path.join("output", slug, lang)
        # helpers mínimos
        def ensure_html(s: str) -> str:
            if "<" in s and ">" in s: return s
            blocks = [f"<p>{line.strip()}</p>" for line in s.split("\n\n") if line.strip()]
            return "<html><body>" + "\n".join(blocks) + "</body></html>"

        attachments = []  # SIEMPRE define por adelantado
        images = []       # SIEMPRE define por adelantado

        if eml_b64:
            print("[DEBUG] usando flujo EML; eml_base64 len:", len(eml_b64))
            try:
                eml_bytes = base64.b64decode(eml_b64, validate=True)
                print("[DEBUG] eml_bytes len:", len(eml_bytes))
            except Exception as e:
                import traceback; traceback.print_exc()
                return jsonify({"ok": False, "where": "base64", "error": str(e)}), 400

            try:
                extracted = extract_html_inline_and_attachments_from_eml_bytes(
                    eml_bytes, slug, lang, append_unreferenced_images=True
                )
            except Exception as e:
                import traceback; traceback.print_exc()
                return jsonify({"ok": False, "where": "extractor", "error": str(e)}), 400

            if not extracted or not isinstance(extracted, dict):
                return jsonify({"ok": False, "where": "extractor", "error": "Extractor devolvió None o tipo no dict"}), 400
            if not extracted.get("html"):
                return jsonify({"ok": False, "where": "extractor", "error": "Extractor sin parte HTML"}), 400

            html_final  = extracted["html"]
            attachments = extracted.get("attachments", [])
            images      = extracted.get("images", [])
            print("[DEBUG] extractor.debug:", extracted.get("debug", {}))

            put_public_s3(
                f"emails/templates/{slug}/{lang}/original.html",
                html_final.encode("utf-8"),
                "text/html; charset=utf-8",
                cache_seconds=0
            )
        else:
            print("[DEBUG] usando flujo HTML (textarea/archivo .html)")
            html_final = ensure_html(html_in)

            # 1) Si el front te manda binarios en payload.attachments, primero
            #    resuelve CIDs con esos binarios (si hay), y separa "file" sin cid.
            payload_attachments = payload.get("attachments") or []
            if payload_attachments:
                html_final, resolved_info = resolve_cid_with_attachments(html_final, slug, payload_attachments)
                # resolved_info: {resolved: [...], unresolved: [...]}  (ids cid que sí/no se mapearon)

            # 2) Rehost de <img> http(s)/data:/ruta → S3 (/images)
            html_final, stats = rehost_images_under_template_from_html(html_final, slug)
            print("[DEBUG] rehost stats:", stats)

            # recopila imágenes subidas (ids y urls) del rehost
            images = stats.get("uploaded", [])

            # Si detectamos <img src="cid:..."> no resueltos y NO nos han enviado binarios, devolvemos 400
            if any(s.get("reason") == "cid_in_html_send_eml" for s in (stats.get("skipped") or [])) and not payload_attachments:
                return jsonify({"error": "El HTML contiene imágenes cid:. Sube el .eml (eml_base64) o envía attachments[] con los binarios."}), 400

            # 3) Insertar y subir ficheros extra SIN cid (pdf/mp4/etc.) que vengan en payload.attachments
            #    a s3://.../emails/templates/<slug>/attachments/<filename>
            if payload_attachments:
                html_final, added_files = insert_extra_files_into_html(html_final, slug, payload_attachments)
                # added_files -> [{"filename","content_type","url"}]
                attachments = added_files

           # ========= NUEVO BLOQUE: separar imágenes de firma (< 30 KB) =========
        SIGNATURE_MAX_BYTES = 30 * 1024  # 30 KB

        signature_images = []
        content_images = []

        for img in images or []:
            # Intenta inferir el tamaño en bytes desde distintos posibles campos
            size = img.get("size") or img.get("filesize") or img.get("length") or 0
            try:
                size = int(size)
            except (TypeError, ValueError):
                size = 0  # si no se puede parsear, lo consideramos “desconocido”

            if size and size < SIGNATURE_MAX_BYTES:
                signature_images.append(img)
            else:
                content_images.append(img)

        images = content_images  # solo las de contenido real en `images`
        # ====================================================================
        body_html, signature_html = split_body_and_signature(html_final)

        # 🔹 limpiar imágenes de la firma (quitar fotos grandes / .jpg, etc.)
        signature_html = clean_signature_images(signature_html)

        # si quieres guardar la firma en algún sitio (S3, parcial, etc.), aquí es el sitio.
        # por ahora seguimos como estabas, usando solo el cuerpo para build_framework:
        html_final = body_html

        # Generar y subir la plantilla (template.html/mjml/schema/manifest)
        result = build_framework(
            input_path_or_html=html_final,
            out_dir=out_dir,
            slug=slug,
            lang=lang,
            upload_to_s3=True,
            display_name=name,
            lang_attachments=attachments
        )

        if signature_html:
        # ruta relativa que quieres usar para el manifest
            signature_key = f"emails/templates/{slug}/partials/signature.html"

            os.makedirs(os.path.join("output", slug, "partials"), exist_ok=True)
            with open(os.path.join("output", slug, "partials", "signature.html"), "w", encoding="utf-8") as f:
                f.write(signature_html)

            put_public_s3(
                signature_key,
                signature_html.encode("utf-8"),
                "text/html; charset=utf-8",
                cache_seconds=0
            )

        # añade metadatos
        manifest_path = result.get("manifest")
        if manifest_path and os.path.exists(manifest_path):
            with open(manifest_path, "r", encoding="utf-8") as f:
                m = json.load(f)
        else:
            m = {}

        m["attachments"] = attachments
        m["images_uploaded"] = images
        if signature_html:
            shared = m.setdefault("shared", {})
            partials = shared.setdefault("partials", {})
            partials.setdefault("signature_html", f"emails/templates/{slug}/partials/signature.html")

        with open(os.path.join(out_dir, "manifest.json"), "w", encoding="utf-8") as f:
            json.dump(m, f, indent=2, ensure_ascii=False)

        return jsonify({
            "template_id": str(uuid.uuid4()),
            "name": name,
            "slug": slug,
            "paths": result,
            "attachments": attachments,
            "images": images,
            "signature_html": signature_html
        }), 200


           






        
    # GET
    else:
        
        return render_template('upload_template_email.html')  # Cambiá 'login' por tu vista de inicio o login



@application.route('/update_template_email', methods=['GET', 'POST'])
def update_template_email():
   
    if request.method == "POST":
        # tu lógica de POST
        pass

    raw = list_email_templates()  
    print ("raw:", raw )          # <- lo que tengas ahora
    items = _coerce_items(raw)     
    print ("items:", items)       # <- **forzamos lista/dict serializable**

    return render_template("update_template_email.html", items=items)
    

        





@application.route("/list_templates", methods=["GET"])
def list_email_templates():
    s3 = boto3.client("s3", region_name=AWS_REGION)

    template_type = (request.args.get("type") or "email").strip().lower()
    if template_type not in ("email", "newsletter"):
        return jsonify({"error": "type debe ser 'email' o 'newsletter'"}), 400

    prefix = "emails/templates/" if template_type == "email" else "newsletters/templates/"
    items = []

    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=prefix, Delimiter="/"):
        for cp in page.get("CommonPrefixes", []):
            base = cp["Prefix"]   # ej: emails/templates/mi-slug/
            slug = base.rstrip("/").split("/")[-1]
            manifest_key = f"{base}manifest.json"

            display_name = slug
            languages = []

            try:
                obj = s3.get_object(Bucket=S3_BUCKET, Key=manifest_key)
                man = json.loads(obj["Body"].read())
                display_name = man.get("display_name") or man.get("name") or man.get("slug") or slug
                languages = sorted((man.get("languages") or {}).keys())
            except botocore.exceptions.ClientError as e:
                code = e.response.get("Error", {}).get("Code")
                if code not in ("NoSuchKey", "404", "NotFound"):
                    raise

            items.append({
                "slug": slug,
                "display_name": display_name,
                "languages": languages,
                "type": template_type
            })

    items.sort(key=lambda x: x["display_name"].lower())
    return jsonify(items), 200


def delete_newsletter_db_by_slug(slug):

    creds = get_db_credentials("secretoBC/Mysql")
    dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

    conn = pymysql.connect(
        host=creds["host"],
        user=creds["username"],
        password=creds["password"],
        database=dbname,
        port=int(creds.get("port", 3306)),
        cursorclass=pymysql.cursors.DictCursor
    )

    try:
        cursor = conn.cursor()

        like_path = f"newsletters/templates/{slug}/%"

        cursor.execute(
            "DELETE FROM newsletters WHERE template_s3_path LIKE %s",
            (like_path,)
        )

        conn.commit()

    finally:
        cursor.close()
        conn.close()
    
   



@application.route('/templates/<slug>', methods=['DELETE'])
def delete_template(slug):
    """
    Borra todos los objetos de:
      - emails/templates/<slug>/
      - newsletters/templates/<slug>/
    según ?type=email|newsletter
    """
    template_type = (request.args.get("type") or "email").strip().lower()
    if template_type not in ("email", "newsletter"):
        return jsonify({"error": "type debe ser 'email' o 'newsletter'"}), 400

    base = "emails/templates" if template_type == "email" else "newsletters/templates"
    prefix = f"{base}/{slug.strip('/')}/"

    s3 = get_s3()
    paginator = s3.get_paginator("list_objects_v2")
    deleted_count = 0
    to_delete_batch = []

    try:
        for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=prefix):
            contents = page.get("Contents", [])
            if not contents:
                continue

            for obj in contents:
                to_delete_batch.append({"Key": obj["Key"]})

                if len(to_delete_batch) == 1000:
                    resp = s3.delete_objects(
                        Bucket=S3_BUCKET,
                        Delete={"Objects": to_delete_batch, "Quiet": True}
                    )
                    deleted_count += len(resp.get("Deleted", []))
                    to_delete_batch = []

        if to_delete_batch:
            resp = s3.delete_objects(
                Bucket=S3_BUCKET,
                Delete={"Objects": to_delete_batch, "Quiet": True}
            )
            deleted_count += len(resp.get("Deleted", []))

        if template_type == "newsletter":
            delete_newsletter_db_by_slug(slug)

        return "", 204

    except botocore.exceptions.ClientError as e:
        application.logger.exception("Error borrando en S3: %s", e)
        return jsonify({"error": "Error al borrar en S3"}), 500

        
@application.route("/templates/<slug>/<lang>/preview", methods=["GET"])
def preview_template_lang(slug, lang):
    template_type = (request.args.get("type") or "email").strip().lower()
    if template_type not in ("email", "newsletter"):
        abort(400, description="type debe ser 'email' o 'newsletter'")

    base_prefix = "emails/templates" if template_type == "email" else "newsletters/templates"

    raw_key = f"{base_prefix}/{slug}/{lang}/original.html"
    tpl_key = f"{base_prefix}/{slug}/{lang}/template.html"
    msg_key = f"{base_prefix}/{slug}/{lang}/partials/message.html"
    sig_key = f"{base_prefix}/{slug}/partials/signature.html"
    manifest_key = f"{base_prefix}/{slug}/manifest.json"
    cidmap_key = f"{base_prefix}/{slug}/cid-map.json"

    use_raw_param = request.args.get("raw")
    force_raw = use_raw_param == "1"
    force_template = use_raw_param == "0"

    if force_raw:
        chosen = raw_key
    elif force_template:
        chosen = tpl_key
    else:
        chosen = tpl_key if s3_key_exists(tpl_key) else raw_key

    if not s3_key_exists(chosen):
        abort(404, description=f"No existe preview para {template_type}/{slug}/{lang}")

    if template_type == "newsletter":
        html = s3_get_text(chosen) or ""
        if not html:
            abort(404, description="No existe HTML para esta newsletter")

        resp = Response(html, mimetype="text/html; charset=utf-8")
        resp.headers["Cache-Control"] = "no-store"
        resp.headers["Content-Security-Policy"] = (
            "default-src 'none'; img-src https: data:; style-src 'unsafe-inline'; "
            "font-src https: data:; frame-ancestors 'self';"
        )
        return resp

    print("[DEBUG] preview type:", template_type, "raw:", use_raw_param, "chosen:", chosen)

    def _load_manifest():
        txt = s3_get_text(manifest_key)
        if not txt:
            return {}
        try:
            return json.loads(txt)
        except Exception:
            return {}

    def _extract_inner_html(html_text: str) -> str:
        if not html_text:
            return ""
        soup = BeautifulSoup(html_text, "lxml")
        if soup.body:
            return "".join(str(x) for x in soup.body.contents)
        return html_text

    def _append_attachments(rendered_html: str, manifest: dict, lang_code: str) -> str:
        try:
            lang_node = (manifest.get("languages") or {}).get(lang_code) or {}
            att_list = lang_node.get("attachments") or []
        except Exception:
            att_list = []

        if not att_list:
            return rendered_html

        block = _attachments_html(att_list)
        try:
            soup_prev = BeautifulSoup(rendered_html, "lxml")
            (soup_prev.body or soup_prev).append(BeautifulSoup(block, "lxml"))
            return str(soup_prev)
        except Exception:
            low = rendered_html.lower()
            idx = low.rfind("</body>")
            return (rendered_html[:idx] + block + rendered_html[idx:]) if idx != -1 else (rendered_html + block)

    manifest = _load_manifest()

    # ---------------------------------------------------------
    # RAW MODE: devuelve original.html tal cual + attachments
    # ---------------------------------------------------------
    if chosen == raw_key:
        html = s3_get_text(raw_key) or ""
        html = _append_attachments(html, manifest, lang)

        resp = Response(html, mimetype="text/html; charset=utf-8")
        resp.headers["Cache-Control"] = "no-store"
        resp.headers["Content-Security-Policy"] = (
            "default-src 'none'; img-src https: data:; style-src 'unsafe-inline'; "
            "font-src https: data:; frame-ancestors 'self';"
        )
        return resp

    # ---------------------------------------------------------
    # TEMPLATE MODE
    # ---------------------------------------------------------
    template_html = s3_get_text(tpl_key) or ""
    message_html = s3_get_text(msg_key) or ""
    signature_html = s3_get_text(sig_key) or ""

    if not template_html:
        abort(404, description="No existe template.html")

    safe_message = _extract_inner_html(message_html)
    safe_signature = _extract_inner_html(signature_html)

    tpl_soup = BeautifulSoup(template_html, "lxml")

    # quita imágenes dentro del slot de firma del template base
    for n in tpl_soup.select('[data-slot="signature"] img, [data-slot="signature"] source'):
        n.extract()

    # bloque de imágenes no-logo desde manifest.shared.images
    shared_images = (manifest.get("shared") or {}).get("images") or {}
    non_logo_imgs = [
        meta for _, meta in shared_images.items()
        if isinstance(meta, dict) and not meta.get("is_logo", False)
    ]

    imgs_holder_soup = BeautifulSoup("<div data-composed='images'></div>", "lxml")
    imgs_holder_div = imgs_holder_soup.div

    for meta in non_logo_imgs:
        src = meta.get("url") or meta.get("key")
        if not src:
            continue

        img = imgs_holder_soup.new_tag("img")
        img["src"] = src

        if meta.get("target_w"):
            img["width"] = meta["target_w"]
        if meta.get("target_h"):
            img["height"] = meta["target_h"]

        imgs_holder_div.append(img)

    out = BeautifulSoup("<!doctype html><html><head></head><body></body></html>", "lxml")

    if tpl_soup.head:
        out.head.replace_with(tpl_soup.head)

    # insertar message como nodos reales, no como string interpolado
    msg_soup = BeautifulSoup(safe_message, "lxml")
    message_nodes = list(msg_soup.body.contents) if msg_soup.body else list(msg_soup.contents)

    message_wrapper = out.new_tag("div")
    message_wrapper["data-composed"] = "message"
    for node in message_nodes:
        message_wrapper.append(node)

    # insertar signature como nodos reales
    sig_soup = BeautifulSoup(safe_signature, "lxml")
    signature_nodes = list(sig_soup.body.contents) if sig_soup.body else list(sig_soup.contents)

    signature_wrapper = out.new_tag("div")
    signature_wrapper["data-composed"] = "signature"
    for node in signature_nodes:
        signature_wrapper.append(node)

    out.body.append(message_wrapper)
    out.body.append(imgs_holder_div)
    out.body.append(signature_wrapper)

    rendered = str(out)

    # reemplazar cid:
    cid_map = {}
    cmap_txt = s3_get_text(cidmap_key)
    if cmap_txt:
        try:
            cid_map = json.loads(cmap_txt)
        except Exception:
            cid_map = {}

    rendered = replace_cid_everywhere(rendered, cid_map)
    rendered = fix_relative_imgs(rendered, slug)
    rendered = apply_manifest_images_all(rendered, manifest, lang=lang)
    rendered = enforce_dimensions_from_manifest(rendered, manifest)
    rendered = inject_preview_css(rendered)

    # attachments
    rendered = _append_attachments(rendered, manifest, lang)

    # quitar logos duplicados fuera de la firma
    try:
        soup_final = BeautifulSoup(rendered, "lxml")
        sig_block = soup_final.select_one("[data-composed='signature']")

        if sig_block:
            sig_keys_final = _collect_image_keys(BeautifulSoup(str(sig_block), "lxml"))
            shared_images = (manifest.get("shared") or {}).get("images") or {}

            for im in list(soup_final.find_all("img")):
                parent = im
                inside_signature = False
                while parent is not None:
                    if getattr(parent, "attrs", None) and parent.attrs.get("data-composed") == "signature":
                        inside_signature = True
                        break
                    parent = getattr(parent, "parent", None)

                if inside_signature:
                    continue

                key = _norm_src(im.get("src", "") or im.get("srcset", ""))
                if key and key in sig_keys_final:
                    meta = shared_images.get(key.lower()) or {}
                    if meta.get("is_logo"):
                        im.decompose()

        rendered = str(soup_final)
    except Exception as e:
        print("[WARN] dedup firmas:", e)

    soup_out = BeautifulSoup(rendered, "lxml")
    print("[DEBUG] preview img count:", len(soup_out.find_all("img")))
    print("[DEBUG] preview first 10 srcs:", [(i.get("src") or "") for i in soup_out.find_all("img")[:10]])

    resp = Response(rendered, mimetype="text/html; charset=utf-8")
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["Content-Security-Policy"] = (
        "default-src 'none'; img-src https: data:; style-src 'unsafe-inline'; "
        "font-src https: data:; connect-src 'self'; frame-ancestors 'self'; "
        "base-uri 'none'; form-action 'none'; script-src 'none'"
    )
    return resp


# app.py (o donde declares tu Flask app)
from flask import Flask, Response, request, abort
from pathlib import Path

# ------------------ API para editar el cuerpo ------------------
@application.put("/api/templates/<slug>/<lang>/partials/message")
def put_message(slug, lang):
    s3 = get_s3()
    body = request.get_data(as_text=True) or ""
    key = key_message(slug, lang)
    resp = s3.put_object(
        Bucket=S3_BUCKET,
        Key=key,
        Body=body.encode("utf-8"),
        ContentType="text/html; charset=utf-8",
        CacheControl="no-cache",
    )
    etag = (resp.get("ETag") or "").strip('"')
    print("[PUT:S3]", {"bucket": S3_BUCKET, "key": key, "bytes": len(body.encode()), "etag": etag})
    return {"ok": True, "bucket": S3_BUCKET, "key": key, "bytes": len(body.encode()), "etag": etag}, 200

@application.get("/api/templates/<slug>/<lang>/partials/message")
def get_message(slug, lang):
    s3 = get_s3()
    for key in (key_message(slug, lang), key_original(slug, lang)):
        try:
            obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
            print("[GET:S3]", {"key": key})
            return Response(obj["Body"].read().decode("utf-8", "replace"),
                            mimetype="text/html; charset=utf-8")
        except botocore.exceptions.ClientError as e:
            if e.response.get("Error", {}).get("Code") in ("NoSuchKey", "404", "NotFound"):
                continue
            raise
    abort(404, "message.html ni original.html")

# --- API: signature (sin idioma) -------------------------------------------


@application.route("/api/templates/<slug>/partials/signature", methods=["GET"])
def api_get_signature(slug):
    s3 = get_s3()
    key = f"emails/templates/{slug}/partials/signature.html"
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
    except botocore.exceptions.ClientError as e:
        code = e.response.get("Error", {}).get("Code")
        if code in ("NoSuchKey", "404", "NotFound"):
            # si prefieres 200 "" en vez de 404, cambia esto
            abort(404, description="signature.html no existe")
        raise
    txt = obj["Body"].read().decode("utf-8", errors="replace")
    return Response(txt, mimetype="text/html; charset=utf-8")


@application.route("/api/templates/<slug>/partials/signature", methods=["PUT"])
def api_put_signature(slug):
    s3 = get_s3()
    key = f"emails/templates/{slug}/partials/signature.html"
    body = request.get_data(as_text=True) or ""
    # Guarda como HTML; si prefieres text/plain, cambia el ContentType
    s3.put_object(
        Bucket=S3_BUCKET,
        Key=key,
        Body=body.encode("utf-8"),
        ContentType="text/html; charset=utf-8",
        CacheControl="no-store",
    )
    return {"ok": True, "mode": "s3", "bucket": S3_BUCKET, "key": key}, 200


# ------------------ PREVIEW compuesto (solo S3) ------------------

@application.get("/templates/<slug>/<lang>/preview")
def preview(slug, lang):
    # raw=1 -> devuelve template.html tal cual desde S3
    raw = request.args.get("raw") == "1"

    tpl = s3_get_text(key_template(slug, lang))
    if tpl is None:
        abort(404, f"template.html no existe en s3://{S3_BUCKET}/{key_template(slug, lang)}")

    if raw:
        return Response(tpl, mimetype="text/html; charset=utf-8")

    # message con fallback a original
    msg = s3_get_text(key_message(slug, lang))
    if msg is None:
        msg = s3_get_text(key_original(slug, lang)) or ""

    sig = s3_get_text(key_signature(slug)) or ""

    # Inyección simple: ajusta a tu sintaxis de marcadores
    html = tpl
    html = re.sub(r"{{\s*>\s*message\s*}}", msg, html)
    html = re.sub(r"{{\s*>\s*signature\s*}}", sig, html)
    html = html.replace("<!-- MESSAGE -->", msg).replace("<!-- SIGNATURE -->", sig)

    return Response(html, mimetype="text/html; charset=utf-8")



@application.route("/list_s3")
def list_s3():
    s3 = boto3.client("s3", region_name=AWS_REGION)
    prefix = request.args.get("prefix") or ROOT_PREFIX_S3
    if not prefix.endswith("/"):
        prefix += "/"

    kwargs = dict(Bucket=S3_BUCKET, Prefix=prefix, Delimiter="/", MaxKeys=1000)
    token = request.args.get("token")
    if token:
        kwargs["ContinuationToken"] = token

    try:
        resp = s3.list_objects_v2(**kwargs)

        folders = []
        for cp in resp.get("CommonPrefixes", []) or []:
            folders.append({
                "name": cp["Prefix"][len(prefix):].rstrip("/"),
                "prefix": cp["Prefix"]
            })

        files = []
        for obj in resp.get("Contents", []) or []:
            key = obj["Key"]
            if key.endswith("/"):
                continue
            relative = key[len(prefix):]
            if "/" in relative:
                continue  # pertenece a subniveles; ya sale en folders
            files.append({
                "key": relative,
                "size": obj.get("Size", 0),
                "last_modified": obj["LastModified"].isoformat(),
                "url": public_url(key)
            })

        return jsonify({
            "ok": True,
            "prefix": prefix,
            "parent_prefix": parent_of(prefix),
            "folders": folders,
            "files": files,
            "is_truncated": bool(resp.get("IsTruncated")),
            "next_token": resp.get("NextContinuationToken"),
            "error": None
        })

   
    except Exception as e:
        err = {"code": "Unexpected", "message": str(e)}

    # ⚠️ En error: mantenemos el mismo shape para no romper el .map()
    return jsonify({
        "ok": False,
        "prefix": prefix,
        "parent_prefix": parent_of(prefix),
        "folders": [],
        "files": [],
        "is_truncated": False,
        "next_token": None,
        "error": err
    }), 500



import dropbox
from flask import jsonify, request
from pathlib import PurePosixPath

def clamp_to_root(path: str) -> str:
    """Fuerza que el path esté bajo DROPBOX_ROOT; si no, devuelve DROPBOX_ROOT."""
    if not path:
        return ROOT_PREFIX_DROPBOX
    # normaliza: sin espacios y siempre con barra inicial
    p = path.strip()
    if not p.startswith("/"):
        p = "/" + p
    # si intenta salir de la raíz virtual, lo fijamos a la raíz
    if not p.startswith(ROOT_PREFIX_DROPBOX):
        return ROOT_PREFIX_DROPBOX
    return p

def get_parent_path(path: str) -> str | None:
    """Calcula la carpeta padre respetando DROPBOX_ROOT."""
    p = PurePosixPath(path)
    parent = str(p.parent)
    if parent == ".":
        parent = ""  # Dropbox usa "" para raíz real
    # Evitar que suba por encima de DROPBOX_ROOT
    if ROOT_PREFIX_DROPBOX and not parent.startswith(ROOT_PREFIX_DROPBOX):
        return None
    return parent

@application.route("/list_dropbox")
def list_dropbox():
    DROPBOX_TOKEN = get_dropbox_access_token()
    dbx = dropbox.Dropbox(DROPBOX_TOKEN)

    raw_path = request.args.get("path", "")
    path = clamp_to_root(raw_path)

    try:
        res = dbx.files_list_folder(path, recursive=False)
    except dropbox.exceptions.ApiError as e:
        return jsonify({
            "ok": False,
            "path": path,
            "parent_path": get_parent_path(path),
            "folders": [],
            "files": [],
            "root_path": ROOT_PREFIX_DROPBOX,
            "error": {"code": "DropboxError", "message": str(e)}
        }), 400

    folders, files = [], []
    for entry in res.entries:
        if isinstance(entry, dropbox.files.FileMetadata):
            tmp = dbx.files_get_temporary_link(entry.path_lower)
            files.append({
                "name": entry.name,
                "path": entry.path_display,
                "size": entry.size,
                "client_modified": entry.client_modified.isoformat(),
                "url": tmp.link
            })
        elif isinstance(entry, dropbox.files.FolderMetadata):
            folders.append({
                "name": entry.name,
                "path": entry.path_display
            })

    return jsonify({
        "ok": True,
        "path": path,
        "parent_path": get_parent_path(path),  # será None cuando estés en /1
        "folders": folders,
        "files": files,
        "error": None
    })

# --- Backend: copiar desde Dropbox a S3 ---


@application.post("/dbx_to_s3")
def dbx_to_s3():
    try:
        data = request.get_json(force=True)
        dbx_path = data.get("dbx_path")
        s3_key   = data.get("s3_key")
        if not dbx_path or not s3_key:
            return jsonify({"ok": False, "error": {"message": "Parámetros requeridos: dbx_path y s3_key"}}), 400

        # 1) Descargar bytes desde Dropbox
        DROPBOX_TOKEN = get_dropbox_access_token()
        dbx = dropbox.Dropbox(DROPBOX_TOKEN)
        

        md, resp = dbx.files_download(dbx_path)
        body = resp.content




        
        

        s3 = boto3.client("s3", region_name=AWS_REGION)

        md5_b64 = base64.b64encode(hashlib.md5(body).digest()).decode("ascii")

        content_type = mimetypes.guess_type(s3_key)[0] or "application/octet-stream"

        s3.put_object(
            Bucket=S3_BUCKET,
            Key=s3_key,
            Body=body,
            ContentType=content_type,
            ContentMD5=md5_b64,
            CacheControl="no-cache, no-store, must-revalidate",
            Expires=0,
        )

        # 3. Actualizar manifest SOLO para esta imagen
        try:
            slug = s3_key.split("/")[2]  # emails/templates/<slug>/...
            updated_manifest = update_manifest_for_key(slug, s3_key)
        except Exception as e:
            # Si algo falla, reconstruye todo
            updated_manifest = update_manifest(slug)

        # 4. Devolver datos útiles al front (etag y last_modified actualizados)
        img_name = s3_key.rsplit("/", 1)[-1]
        img_data = (
            updated_manifest
            .get("shared", {})
            .get("images", {})
            .get(img_name, {})
        )

        return jsonify({
            "ok": True,
            "s3_key": s3_key,
            "etag": img_data.get("etag"),
            "last_modified": img_data.get("last_modified"),
            "url": img_data.get("url"),
        })

      

    except dropbox.exceptions.ApiError as e:
        return jsonify({"ok": False, "error": {"message": f"Dropbox error: {e}"}}), 400
    except boto3.exceptions.Boto3Error as e:
        return jsonify({"ok": False, "error": {"message": f"S3 error: {e}"}}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": {"message": str(e)}}), 500


@application.route('/upload_files_s3', methods=['GET', 'POST'])
def upload_files_s3():
    if request.method == "POST":
        # Procesar la subida del archivo
        pass

    # GET
    else:
        
        return render_template('upload_files_s3.html')  # Cambiá 'login' por tu vista de inicio o login

# routes_campaigns.py


@application.route("/campanas_prospects")
def campanas_prospects():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    ctype = request.args.get("type", "").strip()

    query = (
        db.session.query(
            Campaign,
            func.count(CampaignRecipient.id).label("total_targets")
        )
        .outerjoin(
            CampaignRecipient,
            db.and_(
                CampaignRecipient.campaign_id == Campaign.id,
                CampaignRecipient.entity_kind == "prospect"
            )
        )
        .group_by(Campaign.id)
    )
    if q:
        query = query.filter(Campaign.name.ilike(f"%{q}%"))
    if status:
        query = query.filter(Campaign.status == status)
    if ctype:
        query = query.filter(Campaign.campaign_type == ctype)

    targets = (
        db.session.query(
            ProspectTarget.id,
            ProspectTarget.name.label("nombre_target"),
            ProspectTarget.created_at,
            func.count(ProspectTargetItem.prospect_id).label("total_leads")
        )
        .outerjoin(ProspectTargetItem, ProspectTargetItem.target_id == ProspectTarget.id)
        .group_by(ProspectTarget.id, ProspectTarget.name, ProspectTarget.created_at)
        .order_by(ProspectTarget.name.asc())
        .all()
    )

    query = query.filter(Campaign.status != "sent")

    rows = query.order_by(Campaign.created_at.desc()).all()

    return render_template(
        "campaigns_list.html",
        rows=rows,
        q=q,
        status=status,
        ctype=ctype,
        targets=targets,
        entity_kind="prospect"
        
    )
@application.route("/campanas_leads")
def campanas_leads():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    ctype = request.args.get("type", "").strip()

    query = (
        db.session.query(
            Campaign,
            func.count(CampaignRecipient.id).label("total_targets")
        )
        .outerjoin(
            CampaignRecipient,
            db.and_(
                CampaignRecipient.campaign_id == Campaign.id,
                CampaignRecipient.entity_kind == "lead"
            )
        )
        .group_by(Campaign.id)
    )
    if q:
        query = query.filter(Campaign.name.ilike(f"%{q}%"))
    if status:
        query = query.filter(Campaign.status == status)
    if ctype:
        query = query.filter(Campaign.campaign_type == ctype)

    targets = (
        db.session.query(
            LeadTarget.id,
            LeadTarget.nombre_target,
            LeadTarget.created_at,
            func.count(LeadTargetItem.lead_id).label("total_leads")
        )
        .outerjoin(LeadTargetItem, LeadTargetItem.target_id == LeadTarget.id)
        .group_by(LeadTarget.id, LeadTarget.nombre_target, LeadTarget.created_at)
        .order_by(LeadTarget.nombre_target.asc())
        .all()
    )

    query = query.filter(Campaign.status != 'sent')

    rows = query.order_by(Campaign.created_at.desc()).all()

    return render_template(
        "campaigns_list.html",
        rows=rows,
        q=q,
        status=status,
        ctype=ctype,
        targets=targets,
        entity_kind="lead"
    )
@application.route("/campaigns_history")
def campaigns_history():
    q = request.args.get("q", "").strip()
    entity_kind = request.args.get("entity_kind", "").strip()  # lead / prospect / vacío

    NewsletterES = aliased(Newsletter)
    NewsletterEN = aliased(Newsletter)

    query = (
        db.session.query(
            LeadCampaignHistory.campaign_id,
            LeadCampaignHistory.campaign_name,
            func.max(LeadCampaignHistory.origen).label("origen"),
            func.max(Campaign.idioma).label("idioma"),

            func.max(Campaign.subject_es).label("subject_es"),
            func.max(Campaign.subject_en).label("subject_en"),

            func.max(NewsletterES.name).label("newsletter_es_name"),
            func.max(NewsletterES.template_s3_path).label("newsletter_es_path"),

            func.max(NewsletterEN.name).label("newsletter_en_name"),
            func.max(NewsletterEN.template_s3_path).label("newsletter_en_path"),

            func.count(LeadCampaignHistory.id).label("total_targets"),
            func.max(LeadCampaignHistory.sent_at).label("sent_at")
        )
        .outerjoin(Campaign, Campaign.id == LeadCampaignHistory.campaign_id)
        .outerjoin(NewsletterES, NewsletterES.id == Campaign.newsletter_es_id)
        .outerjoin(NewsletterEN, NewsletterEN.id == Campaign.newsletter_en_id)
        .group_by(
            LeadCampaignHistory.campaign_id,
            LeadCampaignHistory.campaign_name
        )
    )

    if q:
        query = query.filter(
            LeadCampaignHistory.campaign_name.ilike(f"%{q}%")
        )

    if entity_kind:
        query = query.filter(
            LeadCampaignHistory.entity_kind == entity_kind
        )

    rows = query.order_by(
        func.max(LeadCampaignHistory.sent_at).desc()
    ).all()

    return render_template(
        "campaigns_history.html",
        rows=rows,
        q=q,
        entity_kind=entity_kind,
        s3_bucket=S3_BUCKET
    )
@application.route("/preview-newsletter")
def preview_newsletter():
    path = request.args.get("path")

    if not path:
        return "No path", 400

    url = f"https://{S3_BUCKET}.s3.amazonaws.com/{path}"

    return redirect(url)


@application.route("/campanas/nueva", methods=["GET","POST"])
def campaign_new():
    newsletters = Newsletter.query.order_by(Newsletter.created_at.desc()).all()

    print("[DEBUG] GET campanas/nueva - newsletters:", newsletters)

    if request.method == "POST":
        entity_kind = request.form.get("entity_kind", "").strip()
    else:
        entity_kind = request.args.get("entity_kind", "").strip()

    
    print("[DEBUG] campanas/nueva - entity_kind:", entity_kind)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        campaign_type = request.form.get("campaign_type", "").strip()

        newsletter_es_id = request.form.get("newsletter_es_id") or None
        newsletter_en_id = request.form.get("newsletter_en_id") or None

        sender = request.form.get("sender", "").strip()
        reply_to = request.form.get("reply_to", "").strip()

        subject_es = request.form.get("subject_es", "").strip()
        subject_en = request.form.get("subject_en", "").strip()

        idioma = request.form.get("idioma", "es").strip()

        print("POST recibido con datos:", {
            "name": name,
            "campaign_type": campaign_type,
            "newsletter_es_id": newsletter_es_id,
            "newsletter_en_id": newsletter_en_id,
            "sender": sender,
            "reply_to": reply_to,
            "subject_es": subject_es,
            "subject_en": subject_en,
            "idioma": idioma
        })

        if not name:
            flash("El nombre es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

        if campaign_type not in ("emailing", "newsletter"):
            flash("Tipo de campaña inválido", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

        if not sender:
            flash("El sender es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

        if idioma == "es" and not subject_es:
            flash("El subject en español es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

        if idioma == "en" and not subject_en:
            flash("El subject en inglés es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

        if idioma == "both" and (not subject_es or not subject_en):
            flash("Debes indicar subject en español e inglés", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

        if campaign_type == "newsletter":
            if idioma == "es" and not newsletter_es_id:
                flash("Debes seleccionar la newsletter en español", "error")
                return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

            if idioma == "en" and not newsletter_en_id:
                flash("Debes seleccionar la newsletter en inglés", "error")
                return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

            if idioma == "both" and (not newsletter_es_id or not newsletter_en_id):
                flash("Debes seleccionar la newsletter en español e inglés", "error")
                return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

        if campaign_type == "emailing":
            newsletter_es_id = None
            newsletter_en_id = None

        try:
            item = Campaign(
                name=name,
                campaign_type=campaign_type,
                newsletter_es_id=int(newsletter_es_id) if newsletter_es_id else None,
                newsletter_en_id=int(newsletter_en_id) if newsletter_en_id else None,
                sender=sender,
                reply_to=reply_to or None,
                subject_es=subject_es or None,
                subject_en=subject_en or None,
                idioma=idioma,
                status="draft",
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc)
            )

            db.session.add(item)
            db.session.commit()

            flash("Campaña creada", "success")
            endpoint = "campanas_leads" if entity_kind == "lead" else "campanas_prospects"
            return redirect(url_for(endpoint))

        except Exception as e:
            db.session.rollback()
            print("[ERROR] creando campaña:", e)
            flash(f"Error al crear la campaña: {e}", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

    return render_template("campaigns_form.html", newsletters=newsletters, item=None, entity_kind=entity_kind)

@application.route("/campaigns_history/<int:cid>")
def campaigns_history_detail(cid):

    rows = (
        LeadCampaignHistory.query
        .filter(LeadCampaignHistory.campaign_id == cid)
        .order_by(LeadCampaignHistory.sent_at.desc())
        .all()
    )

    # 👇 sacamos nombre y fecha
    campaign_name = rows[0].campaign_name if rows else ""
    sent_at = rows[0].sent_at if rows else None

    return render_template(
        "campaigns_history_detail.html",
        rows=rows,
        cid=cid,
        campaign_name=campaign_name,
        sent_at=sent_at
    )


@application.route("/campanas/<int:cid>/editar", methods=["GET", "POST"])
def campaign_edit(cid):
    item = Campaign.query.get_or_404(cid)
    newsletters = Newsletter.query.order_by(Newsletter.created_at.desc()).all()
    entity_kind = request.values.get("entity_kind", "").strip()

    if request.method == "POST":
        print("POST recibido")
        print(request.form)

        item.name = request.form.get("name", "").strip()
        item.campaign_type = request.form.get("campaign_type", "").strip()
        item.status = request.form.get("status", item.status)

        item.subject_es = request.form.get("subject_es", "").strip()
        item.subject_en = request.form.get("subject_en", "").strip()

        item.sender = request.form.get("sender", "").strip()
        item.reply_to = request.form.get("reply_to", "").strip() or None
        item.idioma = request.form.get("idioma", "es").strip()

        newsletter_es_id = request.form.get("newsletter_es_id") or None
        newsletter_en_id = request.form.get("newsletter_en_id") or None

        if not item.name:
            flash("El nombre es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

        if item.campaign_type not in ("emailing", "newsletter"):
            flash("Tipo de campaña inválido", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

        if not item.sender:
            flash("El sender es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

        if item.idioma not in ("es", "en", "both"):
            flash("Idioma inválido", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

        if item.idioma == "es" and not item.subject_es:
            flash("El subject en español es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

        if item.idioma == "en" and not item.subject_en:
            flash("El subject en inglés es obligatorio", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

        if item.idioma == "both" and (not item.subject_es or not item.subject_en):
            flash("Debes indicar subject en español e inglés", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

        item.newsletter_es_id = None
        item.newsletter_en_id = None

        if item.campaign_type == "newsletter":
            if item.idioma == "es":
                if not newsletter_es_id:
                    flash("Debes seleccionar una newsletter en español", "error")
                    return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)
                item.newsletter_es_id = int(newsletter_es_id)

            elif item.idioma == "en":
                if not newsletter_en_id:
                    flash("Debes seleccionar una newsletter en inglés", "error")
                    return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)
                item.newsletter_en_id = int(newsletter_en_id)

            elif item.idioma == "both":
                if not newsletter_es_id or not newsletter_en_id:
                    flash("Debes seleccionar una newsletter en español y otra en inglés", "error")
                    return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)
                item.newsletter_es_id = int(newsletter_es_id)
                item.newsletter_en_id = int(newsletter_en_id)

        print("Guardando idioma:", item.idioma)
        print("Guardando newsletter_es_id:", item.newsletter_es_id)
        print("Guardando newsletter_en_id:", item.newsletter_en_id)

        try:
            db.session.commit()
            flash("Campaña actualizada", "success")

            if entity_kind == "lead":
                return redirect(url_for("campanas_leads"))
            elif entity_kind == "prospect":
                return redirect(url_for("campanas_prospects"))
            else:
                return redirect(url_for("campanas_leads"))

        except Exception as e:
            db.session.rollback()
            print("ERROR AL GUARDAR:", e)
            flash(f"Error al guardar: {e}", "error")
            return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)

    return render_template("campaigns_form.html", newsletters=newsletters, item=item, entity_kind=entity_kind)
from flask import request, jsonify
import boto3

@application.route("/campanas/<int:cid>/send-test", methods=["POST"])
def campaign_send_test(cid):
    campaign = Campaign.query.get_or_404(cid)

    data = request.get_json(silent=True) or {}
    test_email = (data.get("test_email") or "").strip()

    if not test_email:
        return jsonify({"error": "Debes indicar un email de prueba"}), 400

    try:
        # Elegir newsletter y subject según idioma
        if campaign.idioma == "es":
            if not campaign.newsletter_es_id:
                return jsonify({"error": "La campaña no tiene newsletter en español"}), 400
            newsletter =db.session.get(Newsletter,campaign.newsletter_es_id)
            subject = campaign.subject_es

        elif campaign.idioma == "en":
            if not campaign.newsletter_en_id:
                return jsonify({"error": "La campaña no tiene newsletter en inglés"}), 400
            newsletter = db.session.get(Newsletter,campaign.newsletter_en_id)
            subject = campaign.subject_en

        else:
            # Para prueba, puedes enviar la ES por defecto o decidir otra lógica
            if not campaign.newsletter_es_id:
                return jsonify({"error": "La campaña no tiene newsletter ES para prueba"}), 400
            newsletter = db.session.get(Newsletter,campaign.newsletter_es_id)
            subject = campaign.subject_es

        if not newsletter:
            return jsonify({"error": "Newsletter no encontrada"}), 400

        html = load_newsletter_html(newsletter.template_s3_path)

        
        #unsubscribe_url = f"http://127.0.0.1:8000/unsubscribe?token=preview-demo"

        unsubscribe_url = f"https://api.ledpadel.com/unsubscribe?token=preview-demo"
        final_html = build_final_email_html(
            html,
            lang=campaign.idioma,
            unsubscribe_url=unsubscribe_url
        )

        print(f"[DEBUG] Enviando email de prueba a {test_email} con replay_to '{campaign.reply_to}' con subject '{subject}' y sender '{campaign.sender}'")

        response=send_email_ses(
            to_email=test_email,
            subject=subject,
            html=final_html,
            sender=campaign.sender,
            reply_to=campaign.reply_to
        )
       
        print("SES response:", response)
        return jsonify({
            "ok": True,
            "message": f"Se ha enviado la prueba a {test_email}"
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500



def parse_ses_datetime(value):
    if not value:
        return datetime.now(timezone.utc).replace(tzinfo=None)

    try:
        return datetime.fromisoformat(
            value.replace("Z", "+00:00")
        ).replace(tzinfo=None)
    except (TypeError, ValueError):
        return datetime.now(timezone.utc).replace(tzinfo=None)





@application.route("/aws/ses/notifications", methods=["POST"])
def ses_notifications():

    print("********** ENTRA EN SES NOTIFICATIONS **********")
    print(request.get_data(as_text=True))

    payload = request.get_json(force=True)

    print("ENTRA EN SES NOTIFICATIONS")
    print(payload)

    if payload.get("Type") == "SubscriptionConfirmation":
        subscribe_url = payload.get("SubscribeURL")

        if subscribe_url:
            requests.get(subscribe_url, timeout=10)

        return jsonify({"ok": True})

    if payload.get("Type") != "Notification":
        return jsonify({"ok": True})

    try:
        message = json.loads(payload.get("Message", "{}"))
    except (json.JSONDecodeError, TypeError):
        return jsonify({"ok": True})

    notification_type = message.get("notificationType")
    mail = message.get("mail", {})
    ses_message_id = mail.get("messageId")

    if not ses_message_id:
        return jsonify({"ok": True})

    recipient = CampaignRecipient.query.filter_by(
        ses_message_id=ses_message_id
    ).first()

    if not recipient:
        print(
            "No encuentro recipient para ses_message_id:",
            ses_message_id
        )
        return jsonify({"ok": True})

    if notification_type == "Bounce":
        bounce = message.get("bounce", {})

        bounce_type = bounce.get("bounceType")
        bounce_subtype = bounce.get("bounceSubType")

        bounce_timestamp = parse_ses_datetime(
            bounce.get("timestamp")
        )

        bounced_recipients = bounce.get(
            "bouncedRecipients",
            []
        )

        diagnostic = ""

        if bounced_recipients:
            diagnostic = (
                bounced_recipients[0]
                .get("diagnosticCode", "")
            )

        # Evita incrementar dos veces si SNS reenvía
        # la misma notificación.
        ya_estaba_rebotado = (
            recipient.send_status == "bounced"
            or recipient.bounced_at is not None
        )

        recipient.bounced_at = bounce_timestamp
        recipient.send_status = "bounced"

        recipient.bounce_type = bounce_type
        recipient.bounce_subtype = bounce_subtype
        recipient.bounce_diagnostic = diagnostic or None

        # Mantén también error_message para compatibilidad
        recipient.error_message = (
            diagnostic[:500] if diagnostic else None
        )

        if not ya_estaba_rebotado:
            contacto = None

            if (
                recipient.entity_kind == "lead"
                and recipient.entity_id
            ):
                contacto = db.session.get(
                    LeadForm,
                    recipient.entity_id
                )

            elif (
                recipient.entity_kind == "prospect"
                and recipient.entity_id
            ):
                contacto = db.session.get(
                    ProspectsIA,
                    recipient.entity_id
                )

            if contacto:
                actualizar_estado_rebote_contacto(
                    contacto=contacto,
                    bounce_type=bounce_type,
                    bounce_subtype=bounce_subtype,
                    bounce_timestamp=bounce_timestamp,
                    diagnostic=diagnostic
                )

    elif notification_type == "Delivery":
        delivery = message.get("delivery", {})

        recipient.delivered_at = parse_ses_datetime(
            delivery.get("timestamp")
        )

        recipient.send_status = "delivered"

    elif notification_type == "Complaint":
        complaint = message.get("complaint", {})

        recipient.complained_at = parse_ses_datetime(
            complaint.get("timestamp")
        )

        # Ojo: este valor debe existir en el Enum
        recipient.send_status = "complained"
        recipient.error_message = "Complaint"

        contacto = None

        if (
            recipient.entity_kind == "lead"
            and recipient.entity_id
        ):
            contacto = db.session.get(
                LeadForm,
                recipient.entity_id
            )

        elif (
            recipient.entity_kind == "prospect"
            and recipient.entity_id
        ):
            contacto = db.session.get(
                ProspectsIA,
                recipient.entity_id
            )

        if contacto:
            contacto.email_suppressed = True
            contacto.email_suppressed_at = (
                recipient.complained_at
            )
            contacto.email_suppressed_reason = (
                "Spam complaint"
            )

    db.session.commit()

    print(
        "SES notification processed for recipient "
        f"{recipient.id}: "
        f"type={notification_type}, "
        f"status={recipient.send_status}"
    )

    return jsonify({"ok": True})

@application.route("/campanas/<int:cid>/send", methods=["POST"])
def campaign_send(cid):
    campaign = Campaign.query.get_or_404(cid)
    entity_kind = request.form.get("entity_kind", request.args.get("entity_kind", "")).strip()

    if campaign.status not in ("draft", "ready"):
        flash("La campaña no se puede enviar en este estado", "error")

        if entity_kind == "lead":
            return redirect(url_for("campaign_review", cid=cid, entity_kind=entity_kind))
        elif entity_kind == "prospect":
            return redirect(url_for("campaign_review", cid=cid, entity_kind=entity_kind))
        else:
            return redirect(url_for("campaign_review", cid=cid))

    try:
        sent, failed = send_campaign_batch(cid)

        if sent > 0 and failed == 0:
            campaign.status = "sent"
        elif sent > 0 and failed > 0:
            campaign.status = "partial"
        else:
            campaign.status = "failed"

        db.session.commit()
        flash(f"Campaña enviada. Enviados: {sent}. Fallidos: {failed}.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error enviando campaña: {e}", "error")

    if entity_kind == "lead":
        return redirect(url_for("campaign_review", cid=cid, entity_kind=entity_kind))
    elif entity_kind == "prospect":
        return redirect(url_for("campaign_review", cid=cid, entity_kind=entity_kind))
    else:
        return redirect(url_for("campaign_review", cid=cid))
    




@application.route("/campanas/<int:cid>/generar_target_review")
def campaign_generate_target_review(cid):
    conn = None
    try:
        creds = get_db_credentials("secretoBC/Mysql")
        dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

        conn = pymysql.connect(
            host=creds["host"],
            user=creds["username"],
            password=creds["password"],
            database=dbname,
            port=int(creds.get("port", 3306)),
            cursorclass=pymysql.cursors.DictCursor
        )

        sql = """
            INSERT IGNORE INTO campaign_recipients
            (campaign_id, lead_id, email, pais, idioma, origen, tipo_lead, estado, unsubscribe_token, tracking_id, created_at)

            SELECT
                %s,
                id,
                LOWER(TRIM(email)),
                pais,
                idioma,
                'Leads Oferta',
                tipo_lead,
                estado,
                SHA2(UUID(),256),
                UUID(),
                NOW()

            FROM lead_forms
            WHERE email IS NOT NULL
            AND email <> ''
            AND (unsubscribed IS NULL OR unsubscribed = '0' OR unsubscribed = 0)

            GROUP BY LOWER(TRIM(email))
            """
        with conn.cursor() as cur:
            rows = cur.execute(sql, (cid,))
        conn.commit()

        flash(f"Target generado correctamente ({rows} filas insertadas/intentadas).", "success")

    except Exception as e:
        print(f"Error al generar target para campaña {cid}: {e}")
        flash("Error al generar target: " + str(e), "error")

    finally:
        if conn:
            conn.close()

    return redirect(url_for("campanas"))






@application.route("/campanas/<int:cid>/review")
def campaign_review(cid):
    campaign = Campaign.query.get_or_404(cid)
    entity_kind = request.args.get("entity_kind", "").strip()

    newsletter_es = campaign.newsletter_es
    newsletter_en = campaign.newsletter_en

    total_sql = """
        SELECT COUNT(*)
        FROM campaign_recipients
        WHERE campaign_id = :cid
    """

    selected_sql = """
        SELECT COUNT(*)
        FROM campaign_recipients
        WHERE campaign_id = :cid
          AND seleccionado = 1
    """

    params = {"cid": cid}

    if entity_kind:
        total_sql += " AND entity_kind = :entity_kind"
        selected_sql += " AND entity_kind = :entity_kind"
        params["entity_kind"] = entity_kind

    total = db.session.execute(
        db.text(total_sql),
        params
    ).scalar()

    selected = db.session.execute(
        db.text(selected_sql),
        params
    ).scalar()

    return render_template(
        "campaigns_review.html",
        campaign=campaign,
        newsletter_es=newsletter_es,
        newsletter_en=newsletter_en,
        total=total,
        selected=selected,
        entity_kind=entity_kind
    )

@application.route("/campanas/<int:cid>/borrar", methods=["POST"])
def campaign_delete(cid):
    item = Campaign.query.get_or_404(cid)
    entity_kind = request.args.get("entity_kind", "").strip()

    try:
        db.session.delete(item)
        db.session.commit()
        flash("Campaña borrada correctamente", "success")
    except Exception as e:
        db.session.rollback()
        print("[ERROR] borrando campaña:", e)
        flash(f"Error al borrar la campaña: {e}", "error")

    if entity_kind == "lead":
        return redirect(url_for("campanas_leads"))
    elif entity_kind == "prospect":
        return redirect(url_for("campanas_prospects"))
    else:
        return redirect(url_for("campanas_leads"))



@application.route("/campanas/<int:cid>/generar_target/<int:target_id>")
def campaign_generate_target(cid, target_id):
    conn = None
    entity_kind = request.args.get("entity_kind", "").strip()

    try:
        creds = get_db_credentials("secretoBC/Mysql")
        dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

        conn = pymysql.connect(
            host=creds["host"],
            user=creds["username"],
            password=creds["password"],
            database=dbname,
            port=int(creds.get("port", 3306)),
            cursorclass=pymysql.cursors.DictCursor
        )

        sql_delete = """
            DELETE FROM campaign_recipients
            WHERE campaign_id = %s
              AND entity_kind = %s
        """

        if entity_kind == "lead":
            sql_insert = """
                INSERT INTO campaign_recipients
                (
                    campaign_id,
                    entity_kind,
                    entity_id,
                    email,
                    pais,
                    idioma,
                    origen,
                    tipo_lead,
                    estado,
                    unsubscribe_token,
                    tracking_id,
                    created_at
                )
                SELECT
                    %s,
                    'lead',
                    lf.id,
                    LOWER(TRIM(lf.email)),
                    lf.pais,
                    lf.idioma,
                    lt.nombre_target,
                    lf.tipo_lead,
                    lf.estado,
                    SHA2(UUID(), 256),
                    UUID(),
                    NOW()
                FROM lead_target_items lti
                INNER JOIN lead_targets lt
                    ON lt.id = lti.target_id
                INNER JOIN lead_forms lf
                    ON lf.id = lti.lead_id
                INNER JOIN (
                    SELECT
                        MIN(lf2.id) AS selected_id
                    FROM lead_target_items lti2
                    INNER JOIN lead_forms lf2
                        ON lf2.id = lti2.lead_id
                    WHERE lti2.target_id = %s
                      AND lf2.email IS NOT NULL
                      AND TRIM(lf2.email) <> ''
                      AND (lf2.unsubscribed IS NULL OR lf2.unsubscribed = 0 OR lf2.unsubscribed = '0')
                    GROUP BY LOWER(TRIM(lf2.email))
                ) dedup
                    ON dedup.selected_id = lf.id
                WHERE lti.target_id = %s
            """

        elif entity_kind == "prospect":
            sql_insert = """
                INSERT INTO campaign_recipients
                (
                    campaign_id,
                    entity_kind,
                    entity_id,
                    email,
                    pais,
                    idioma,
                    origen,
                    tipo_lead,
                    estado,
                    unsubscribe_token,
                    tracking_id,
                    created_at
                )
                SELECT
                    %s,
                    'prospect',
                    pf.id,
                    LOWER(TRIM(pf.email)),
                    pf.pais,
                    pf.idioma,
                    pt.name,
                    pf.tipo,
                    pf.estado,
                    SHA2(UUID(), 256),
                    UUID(),
                    NOW()
                FROM prospect_target_items pti
                INNER JOIN prospect_targets pt
                    ON pt.id = pti.target_id
                INNER JOIN prospects_IA pf
                    ON pf.id = pti.prospect_id
                INNER JOIN (
                    SELECT
                        MIN(pf2.id) AS selected_id
                    FROM prospect_target_items pti2
                    INNER JOIN prospects_IA pf2
                        ON pf2.id = pti2.prospect_id
                    WHERE pti2.target_id = %s
                    AND pf2.email IS NOT NULL
                    AND TRIM(pf2.email) <> ''
                    AND (pf2.unsubscribed IS NULL OR pf2.unsubscribed = 0 OR pf2.unsubscribed = '0')
                    GROUP BY LOWER(TRIM(pf2.email))
                ) dedup
                    ON dedup.selected_id = pf.id
                WHERE pti.target_id = %s
            """
        else:
            flash("Tipo de entidad inválido", "error")
            return redirect(url_for("campanas_leads"))

        with conn.cursor() as cur:
            cur.execute(sql_delete, (cid, entity_kind))
            deleted_rows = cur.rowcount

            cur.execute(sql_insert, (cid, target_id, target_id))
            inserted_rows = cur.rowcount

        conn.commit()

        flash(
            f"Target regenerado correctamente. "
            f"Destinatarios anteriores borrados: {deleted_rows}. "
            f"Nuevos insertados: {inserted_rows}.",
            "success"
        )

    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Error al generar target para campaña {cid}, target {target_id}: {e}")
        flash("Error al generar target: " + str(e), "error")

    finally:
        if conn:
            conn.close()

    if entity_kind == "lead":
        return redirect(url_for("campanas_leads"))
    elif entity_kind == "prospect":
        return redirect(url_for("campanas_prospects"))
    return redirect(url_for("campanas_leads"))


@application.route("/campanas/<int:cid>/targets")
def campaign_targets(cid):
    campaign = Campaign.query.get_or_404(cid)

    entity_kind = request.args.get("entity_kind", "").strip()
    pais = request.args.get("pais", "").strip()
    idioma = request.args.get("idioma", "").strip()
    origen = request.args.get("origen", "").strip()
    estado_envio = request.args.get("estado_envio", "").strip()
    estado_lead = request.args.get("estado_lead", "").strip()
    tipo_lead = request.args.get("tipo_lead", "").strip()

    print(f"[DEBUG] campaign_targets - filtros recibidos: entity_kind='{entity_kind}', pais='{pais}', idioma='{idioma}', origen='{origen}', estado_envio='{estado_envio}', estado_lead='{estado_lead}', tipo_lead='{tipo_lead}'")

    q = CampaignRecipient.query.filter(CampaignRecipient.campaign_id == cid)

    if entity_kind:
        q = q.filter(CampaignRecipient.entity_kind == entity_kind)

    if pais:
        q = q.filter(CampaignRecipient.pais == pais)

    if idioma:
        q = q.filter(CampaignRecipient.idioma == idioma)

    if origen:
        q = q.filter(CampaignRecipient.origen == origen)

    if estado_envio:
        q = q.filter(CampaignRecipient.send_status == estado_envio)

    if estado_lead:
        q = q.filter(CampaignRecipient.estado == estado_lead)

    if tipo_lead:
        q = q.filter(CampaignRecipient.tipo_lead == tipo_lead)

    targets = q.order_by(CampaignRecipient.email).all()

    base_filters = [CampaignRecipient.campaign_id == cid]
    if entity_kind:
        base_filters.append(CampaignRecipient.entity_kind == entity_kind)

    entity_kinds = (
        db.session.query(CampaignRecipient.entity_kind)
        .filter(*base_filters)
        .filter(CampaignRecipient.entity_kind.isnot(None))
        .distinct()
        .order_by(CampaignRecipient.entity_kind)
        .all()
    )
    entity_kinds = [x[0] for x in entity_kinds]

    paises = (
        db.session.query(func.trim(CampaignRecipient.pais))
        .filter(*base_filters)
        .filter(CampaignRecipient.pais.isnot(None))
        .filter(func.trim(CampaignRecipient.pais) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.pais))
        .all()
    )
    paises = [p[0] for p in paises]

    idiomas = (
        db.session.query(func.trim(CampaignRecipient.idioma))
        .filter(*base_filters)
        .filter(CampaignRecipient.idioma.isnot(None))
        .filter(func.trim(CampaignRecipient.idioma) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.idioma))
        .all()
    )
    idiomas = [i[0] for i in idiomas]

    origenes = (
        db.session.query(func.trim(CampaignRecipient.origen))
        .filter(*base_filters)
        .filter(CampaignRecipient.origen.isnot(None))
        .filter(func.trim(CampaignRecipient.origen) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.origen))
        .all()
    )
    origenes = [o[0] for o in origenes]

    estados_envio = (
        db.session.query(CampaignRecipient.send_status)
        .filter(*base_filters)
        .filter(CampaignRecipient.send_status.isnot(None))
        .distinct()
        .order_by(CampaignRecipient.send_status)
        .all()
    )
    estados_envio = [s[0] for s in estados_envio]

    estados_lead = (
        db.session.query(func.trim(CampaignRecipient.estado))
        .filter(*base_filters)
        .filter(CampaignRecipient.estado.isnot(None))
        .filter(func.trim(CampaignRecipient.estado) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.estado))
        .all()
    )
    estados_lead = [e[0] for e in estados_lead]

    tipos_lead = (
        db.session.query(func.trim(CampaignRecipient.tipo_lead))
        .filter(*base_filters)
        .filter(CampaignRecipient.tipo_lead.isnot(None))
        .filter(func.trim(CampaignRecipient.tipo_lead) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.tipo_lead))
        .all()
    )
    tipos_lead = [t[0] for t in tipos_lead]

    return render_template(
        "campaign_targets.html",
        campaign=campaign,
        targets=targets,
        entity_kind=entity_kind,
        entity_kinds=entity_kinds,
        pais=pais,
        idioma=idioma,
        origen=origen,
        estado_envio=estado_envio,
        estado_lead=estado_lead,
        tipo_lead=tipo_lead,
        paises=paises,
        idiomas=idiomas,
        origenes=origenes,
        estados_envio=estados_envio,
        estados_lead=estados_lead,
        tipos_lead=tipos_lead
    )



@application.route("/campanas/<int:cid>/targets/<int:rid>/toggle", methods=["POST"])
def campaign_toggle_target(cid, rid):
    entity_kind = request.args.get("entity_kind", "").strip()

    query = CampaignRecipient.query.filter_by(
        id=rid,
        campaign_id=cid
    )

    if entity_kind:
        query = query.filter(CampaignRecipient.entity_kind == entity_kind)

    target = query.first_or_404()

    new_value = request.form.get("seleccionado", "1")
    target.seleccionado = True if str(new_value) == "1" else False

    db.session.commit()

    return redirect(url_for(
        "campaign_targets",
        cid=cid,
        entity_kind=entity_kind,
        pais=request.args.get("pais", ""),
        idioma=request.args.get("idioma", ""),
        origen=request.args.get("origen", ""),
        estado_envio=request.args.get("estado_envio", ""),
        estado_lead=request.args.get("estado_lead", ""),
        tipo_lead=request.args.get("tipo_lead", "")
    ))

def _filtered_targets_query(cid):
    entity_kind = request.args.get("entity_kind", "").strip()
    pais = request.args.get("pais", "").strip()
    idioma = request.args.get("idioma", "").strip()
    origen = request.args.get("origen", "").strip()
    estado_envio = request.args.get("estado_envio", "").strip()
    estado_lead = request.args.get("estado_lead", "").strip()
    tipo_lead = request.args.get("tipo_lead", "").strip()

    q = CampaignRecipient.query.filter(CampaignRecipient.campaign_id == cid)

    if entity_kind:
        q = q.filter(CampaignRecipient.entity_kind == entity_kind)

    if pais:
        q = q.filter(CampaignRecipient.pais == pais)

    if idioma:
        q = q.filter(CampaignRecipient.idioma == idioma)

    if origen:
        q = q.filter(CampaignRecipient.origen == origen)

    if estado_envio:
        q = q.filter(CampaignRecipient.send_status == estado_envio)

    if estado_lead:
        q = q.filter(CampaignRecipient.estado == estado_lead)

    if tipo_lead:
        q = q.filter(CampaignRecipient.tipo_lead == tipo_lead)

    return q


@application.route("/campanas/<int:cid>/targets/select_all")
def campaign_select_all_filtered(cid):
    q = _filtered_targets_query(cid)
    q.update({"seleccionado": True}, synchronize_session=False)
    db.session.commit()

    return redirect(url_for(
        "campaign_targets",
        cid=cid,
        entity_kind=request.args.get("entity_kind", ""),
        pais=request.args.get("pais", ""),
        idioma=request.args.get("idioma", ""),
        origen=request.args.get("origen", ""),
        estado_envio=request.args.get("estado_envio", ""),
        estado_lead=request.args.get("estado_lead", ""),
        tipo_lead=request.args.get("tipo_lead", "")
    ))

@application.route("/campanas/<int:cid>/targets/unselect_all")
def campaign_unselect_all_filtered(cid):
    q = _filtered_targets_query(cid)
    q.update({"seleccionado": False}, synchronize_session=False)
    db.session.commit()

    return redirect(url_for(
        "campaign_targets",
        cid=cid,
        entity_kind=request.args.get("entity_kind", ""),
        pais=request.args.get("pais", ""),
        idioma=request.args.get("idioma", ""),
        origen=request.args.get("origen", ""),
        estado_envio=request.args.get("estado_envio", ""),
        estado_lead=request.args.get("estado_lead", ""),
        tipo_lead=request.args.get("tipo_lead", "")
    ))






@application.route("/campanas/<int:cid>/targets/seleccionados")
def campaign_view_selected(cid):
    campaign = Campaign.query.get_or_404(cid)

    entity_kind = request.args.get("entity_kind", "").strip()
    pais = request.args.get("pais", "").strip()
    idioma = request.args.get("idioma", "").strip()
    origen = request.args.get("origen", "").strip()
    estado_envio = request.args.get("estado_envio", "").strip()
    estado_lead = request.args.get("estado_lead", "").strip()
    tipo_lead = request.args.get("tipo_lead", "").strip()

    q = CampaignRecipient.query.filter(
        CampaignRecipient.campaign_id == cid,
        CampaignRecipient.seleccionado == True
    )

    if entity_kind:
        q = q.filter(CampaignRecipient.entity_kind == entity_kind)

    if pais:
        q = q.filter(CampaignRecipient.pais == pais)

    if idioma:
        q = q.filter(CampaignRecipient.idioma == idioma)

    if origen:
        q = q.filter(CampaignRecipient.origen == origen)

    if estado_envio:
        q = q.filter(CampaignRecipient.send_status == estado_envio)

    if estado_lead:
        q = q.filter(CampaignRecipient.estado == estado_lead)

    if tipo_lead:
        q = q.filter(CampaignRecipient.tipo_lead == tipo_lead)

    targets = q.order_by(CampaignRecipient.email).all()

    base_filters = [CampaignRecipient.campaign_id == cid]
    if entity_kind:
        base_filters.append(CampaignRecipient.entity_kind == entity_kind)

    paises = [p[0] for p in (
        db.session.query(func.trim(CampaignRecipient.pais))
        .filter(*base_filters)
        .filter(CampaignRecipient.pais.isnot(None))
        .filter(func.trim(CampaignRecipient.pais) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.pais))
        .all()
    )]

    idiomas = [i[0] for i in (
        db.session.query(func.trim(CampaignRecipient.idioma))
        .filter(*base_filters)
        .filter(CampaignRecipient.idioma.isnot(None))
        .filter(func.trim(CampaignRecipient.idioma) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.idioma))
        .all()
    )]

    origenes = [o[0] for o in (
        db.session.query(func.trim(CampaignRecipient.origen))
        .filter(*base_filters)
        .filter(CampaignRecipient.origen.isnot(None))
        .filter(func.trim(CampaignRecipient.origen) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.origen))
        .all()
    )]

    estados_envio = [s[0] for s in (
        db.session.query(CampaignRecipient.send_status)
        .filter(*base_filters)
        .filter(CampaignRecipient.send_status.isnot(None))
        .distinct()
        .order_by(CampaignRecipient.send_status)
        .all()
    )]

    estados_lead = [e[0] for e in (
        db.session.query(func.trim(CampaignRecipient.estado))
        .filter(*base_filters)
        .filter(CampaignRecipient.estado.isnot(None))
        .filter(func.trim(CampaignRecipient.estado) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.estado))
        .all()
    )]

    tipos_lead = [t[0] for t in (
        db.session.query(func.trim(CampaignRecipient.tipo_lead))
        .filter(*base_filters)
        .filter(CampaignRecipient.tipo_lead.isnot(None))
        .filter(func.trim(CampaignRecipient.tipo_lead) != "")
        .distinct()
        .order_by(func.trim(CampaignRecipient.tipo_lead))
        .all()
    )]

    return render_template(
        "campaign_targets.html",
        campaign=campaign,
        targets=targets,
        entity_kind=entity_kind,
        pais=pais,
        idioma=idioma,
        origen=origen,
        estado_envio=estado_envio,
        estado_lead=estado_lead,
        tipo_lead=tipo_lead,
        paises=paises,
        idiomas=idiomas,
        origenes=origenes,
        estados_envio=estados_envio,
        estados_lead=estados_lead,
        tipos_lead=tipos_lead
    )
@application.route("/newsletters/<int:nid>/preview")
def newsletter_preview(nid):

    lang = request.args.get("lang", "es")

    newsletter = Newsletter.query.get_or_404(nid)

    path = newsletter.template_s3_path

    if lang == "en":
        path = path.replace("/es/", "/en/")

    s3 = boto3.client("s3", region_name="eu-north-1")

    obj = s3.get_object(
        Bucket="emailingledpadel",
        Key=path
    )

    html = obj["Body"].read().decode("utf-8")

    return Response(html, mimetype="text/html")



@application.route("/privacy")
def privacy_policy():
    return """
    <h2>Privacy Policy</h2>
    <p>This page will contain our privacy policy.</p>
    """

@application.route("/preferences")
def email_preferences():
    return """
    <h2>Email Preferences</h2>
    <p>This page will allow users to manage email preferences.</p>
    """
@application.route("/unsubscribe")
def unsubscribe():

    token = request.args.get("token")

    print("Token recibido para baja:", token)

    if not token:
        return "Token inválido", 400
    
    if token == 'preview-demo':
        lang =  "en"
        email=''
    else:
        recipient = CampaignRecipient.query.filter_by(
            unsubscribe_token=token
        ).first()

        if not recipient:
            return "Token inválido", 404

        lang = recipient.idioma or "es"
        email=recipient.email

    if lang == "en":
        template = "unsubscribe_en.html"
    else:
        template = "unsubscribe_es.html"

    return render_template(
        template,
        token=token,
        email=email
        )

@application.route("/unsubscribe", methods=["POST"])
def unsubscribe_submit():
    print("Formulario de baja recibido con datos:", request.form)
    token = request.form.get("token")
    print("Token recibido para procesar baja:", token)

    if token == "preview-demo":
        return render_template(
            "unsubscribe_en_OK.html",
            preview=True
        )
    recipient = CampaignRecipient.query.filter_by(
        unsubscribe_token=token
    ).first()

    lang = recipient.idioma if recipient else "es"  
    if not recipient:
        return "Token inválido", 404

    # marcar baja global por email o lead
    lead = LeadForm.query.filter_by(email=recipient.email).first()
    if lead:
        lead.unsubscribed = True
        lead.unsubscribed_at = datetime.now(timezone.utc)

    db.session.commit()

    if lang == "en":
        template = "unsubscribe_en_OK.html"
    else:
        template = "unsubscribe_es_OK.html"

    return render_template(
        template,
        preview=False,
        
        )

@application.route("/email/open")
def email_open():
    tracking_id = request.args.get("id", "").strip()

    print(f"Open recibido. Tracking ID: {tracking_id}")

    if tracking_id:
        recipient = db.session.query(CampaignRecipient).filter_by(
            tracking_id=tracking_id
        ).first()

        if recipient and not recipient.opened_at:
            recipient.opened_at = datetime.now(timezone.utc)
            db.session.commit()

    # Pixel PNG transparente 1x1
    pixel = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO8Z3ZkAAAAASUVORK5CYII="
    )

    return Response(pixel, mimetype="image/png")


@application.route("/email/click")
def email_click():
    try:
        tracking_id = request.args.get("id", "").strip()
        target_url = request.args.get("url", "").strip()

        print(f"Tracking ID: {tracking_id}")
        print(f"Target URL: {target_url}")

        recipient = db.session.query(CampaignRecipient).filter_by(
            tracking_id=tracking_id
        ).first()

        print(f"Recipient encontrado: {recipient}")

        if recipient:
            print(f"Email: {recipient.email}")
            print(f"Click count actual: {recipient.click_count}")

            if not recipient.clicked_at:
                recipient.clicked_at = datetime.now(timezone.utc)

            recipient.click_count = (recipient.click_count or 0) + 1

            print("Antes de commit")
            db.session.commit()
            print("Commit OK")

        print("Redirigiendo...")
        return redirect(target_url, code=302)

    except Exception as e:


        
        import traceback
        traceback.print_exc()
        return str(e), 500


@application.route("/campanas/<int:cid>/send-stream")
def campaign_send_stream(cid):
    entity_kind = request.args.get("entity_kind", "").strip()

    def generate():
        try:
            for item in send_campaign_batch_stream(cid, entity_kind=entity_kind):
                yield f"data: {json.dumps(item)}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no"
        }
    )

@application.route("/campanas/<int:cid>/stats")
def campaign_stats(cid):
    campaign = Campaign.query.get(cid)

    if not campaign:
        hist = LeadCampaignHistory.query.filter_by(campaign_id=cid).first()
        if not hist:
            abort(404)

        campaign = SimpleNamespace(
            id=cid,
            name=hist.campaign_name,
            campaign_name=hist.campaign_name,
            subject_es=None,
            subject_en=None,
            idioma=None
        )

    entity_kind = request.args.get("entity_kind", "").strip()

    stats_query = db.session.query(
        func.count(CampaignRecipient.id).label("total"),
        func.sum(func.if_(CampaignRecipient.send_status == "pending", 1, 0)).label("pending"),
        func.sum(func.if_(CampaignRecipient.sent_at.isnot(None), 1, 0)).label("sent"),
        func.sum(func.if_(CampaignRecipient.delivered_at.isnot(None), 1, 0)).label("delivered"),
        func.sum(func.if_(CampaignRecipient.opened_at.isnot(None), 1, 0)).label("opened"),
        func.sum(func.if_(CampaignRecipient.clicked_at.isnot(None), 1, 0)).label("clicked"),
        func.sum(func.if_(CampaignRecipient.bounced_at.isnot(None), 1, 0)).label("bounced"),
        func.sum(func.if_(CampaignRecipient.complained_at.isnot(None), 1, 0)).label("complained"),
        func.sum(func.if_(CampaignRecipient.send_status == "error", 1, 0)).label("errors"),
        func.coalesce(func.sum(CampaignRecipient.click_count), 0).label("total_clicks")
    ).filter(
        CampaignRecipient.campaign_id == cid
    )
    

    if entity_kind:
        stats_query = stats_query.filter(CampaignRecipient.entity_kind == entity_kind)

    stats = stats_query.first()

    if entity_kind == "lead":
        unsubscribed = (
            db.session.query(func.count(CampaignRecipient.id))
            .join(LeadForm, CampaignRecipient.entity_id == LeadForm.id)
            .filter(
                CampaignRecipient.campaign_id == cid,
                CampaignRecipient.entity_kind == "lead",
                LeadForm.unsubscribed == 1
            )
            .scalar() or 0
        )

    elif entity_kind == "prospect":
        unsubscribed = (
            db.session.query(func.count(CampaignRecipient.id))
            .join(ProspectsIA, CampaignRecipient.entity_id == ProspectsIA.id)
            .filter(
                CampaignRecipient.campaign_id == cid,
                CampaignRecipient.entity_kind == "prospect",
                ProspectsIA.unsubscribed == 1
            )
            .scalar() or 0
        )

    else:
        lead_unsubscribed = (
            db.session.query(func.count(CampaignRecipient.id))
            .join(LeadForm, CampaignRecipient.entity_id == LeadForm.id)
            .filter(
                CampaignRecipient.campaign_id == cid,
                CampaignRecipient.entity_kind == "lead",
                LeadForm.unsubscribed == 1
            )
            .scalar() or 0
        )

        prospect_unsubscribed = (
            db.session.query(func.count(CampaignRecipient.id))
            .join(ProspectsIA, CampaignRecipient.entity_id == ProspectsIA.id)
            .filter(
                CampaignRecipient.campaign_id == cid,
                CampaignRecipient.entity_kind == "prospect",
                ProspectsIA.unsubscribed == 1
            )
            .scalar() or 0
        )
        unsubscribed = lead_unsubscribed + prospect_unsubscribed


    total = stats.total or 0
    pending = stats.pending or 0
    sent = stats.sent or 0
    delivered = stats.delivered or 0
    opened = stats.opened or 0
    clicked = stats.clicked or 0
    bounced = stats.bounced or 0
    errors = stats.errors or 0
    
    complained = stats.complained or 0
    total_clicks = stats.total_clicks or 0

    # De momento, si no tienes unsubscribe en prospects, mejor dejarlo a 0
    print(f"[DEBUG] campaign_stats - unsubscribed: '{unsubscribed}'")

    delivery_rate = round((delivered / sent) * 100, 2) if sent else 0
    open_rate = round((opened / sent) * 100, 2) if sent else 0
    click_rate = round((clicked / sent) * 100, 2) if sent else 0
    ctor_rate = round((clicked / opened) * 100, 2) if opened else 0
    bounce_rate = round((bounced / sent) * 100, 2) if sent else 0
    unsubscribe_rate = round((unsubscribed / sent) * 100, 2) if sent else 0

    by_status_query = db.session.query(
        CampaignRecipient.send_status,
        func.count(CampaignRecipient.id)
    ).filter(
        CampaignRecipient.campaign_id == cid
    )

    if entity_kind:
        by_status_query = by_status_query.filter(CampaignRecipient.entity_kind == entity_kind)

    by_status = by_status_query.group_by(
        CampaignRecipient.send_status
    ).all()

    by_country_query = db.session.query(
        CampaignRecipient.pais,
        func.count(CampaignRecipient.id)
    ).filter(
        CampaignRecipient.campaign_id == cid
    )

    if entity_kind:
        by_country_query = by_country_query.filter(CampaignRecipient.entity_kind == entity_kind)

    by_country = by_country_query.group_by(
        CampaignRecipient.pais
    ).order_by(
        func.count(CampaignRecipient.id).desc()
    ).all()

    
    print(f"[DEBUG] entity_kind = '{entity_kind}'")

    return render_template(
        "campaign_stats.html",
        campaign=campaign,
        entity_kind=entity_kind,
        total=total,
        pending=pending,
        sent=sent,
        delivered=delivered,
        opened=opened,
        clicked=clicked,
        bounced=bounced,
        unsubscribed=unsubscribed,
        errors=errors,
        complained=complained,
        total_clicks=total_clicks,
        delivery_rate=delivery_rate,
        open_rate=open_rate,
        click_rate=click_rate,
        ctor_rate=ctor_rate,
        bounce_rate=bounce_rate,
        by_status=by_status,
        by_country=by_country,
        unsubscribe_rate=unsubscribe_rate
    )


@application.route("/store_prospect_targets", methods=["POST"])
def store_prospect_targets():
    data = request.get_json(silent=True) or {}

    nombre_target = (data.get("nombre_target") or "").strip()
    prospect_ids = data.get("prospect_ids") or []

    if not nombre_target:
        return jsonify({
            "ok": False,
            "error": "nombre_target es obligatorio"
        }), 400

    if not isinstance(prospect_ids, list) or not prospect_ids:
        return jsonify({
            "ok": False,
            "error": "prospect_ids debe ser una lista con al menos un id"
        }), 400

    try:
        prospect_ids = [int(x) for x in prospect_ids]
    except (TypeError, ValueError):
        return jsonify({
            "ok": False,
            "error": "Todos los prospect_ids deben ser numéricos"
        }), 400

    # Quitar duplicados manteniendo el orden
    prospect_ids = list(dict.fromkeys(prospect_ids))

    print(
        f"[INFO] Almacenando prospect target para '{nombre_target}' "
        f"con {len(prospect_ids)} prospectos: {prospect_ids}"
    )

    creds = get_db_credentials("secretoBC/Mysql")
    dbname = "bc_pruebas" if BD == "PRUEBAS" else creds["dbname"]

    conn = pymysql.connect(
        host=creds["host"],
        user=creds["username"],
        password=creds["password"],
        database=dbname,
        port=int(creds.get("port", 3306)),
        autocommit=False
    )

    try:
        with conn.cursor() as cur:
            # 1) Validar existencia, baja y bloqueo de email
            placeholders = ",".join(["%s"] * len(prospect_ids))

            sql_check = f"""
                SELECT
                    id,
                    COALESCE(unsubscribed, 0) AS unsubscribed,
                    COALESCE(email_suppressed, 0) AS email_suppressed,
                    email_suppressed_reason
                FROM prospects_IA
                WHERE id IN ({placeholders})
            """

            cur.execute(sql_check, tuple(prospect_ids))
            rows = cur.fetchall()

            existing_ids = {row[0] for row in rows}

            unsubscribed_ids = {
                row[0]
                for row in rows
                if int(row[1] or 0) == 1
            }

            suppressed_ids = {
                row[0]
                for row in rows
                if int(row[2] or 0) == 1
            }

            suppressed_reasons = {
                row[0]: row[3]
                for row in rows
                if int(row[2] or 0) == 1
            }

            missing_ids = [
                prospect_id
                for prospect_id in prospect_ids
                if prospect_id not in existing_ids
            ]

            if missing_ids:
                conn.rollback()

                return jsonify({
                    "ok": False,
                    "error": "Algunos prospectos no existen",
                    "missing_ids": missing_ids
                }), 400

            # Un prospecto puede estar dado de baja y bloqueado a la vez
            excluded_ids = unsubscribed_ids | suppressed_ids

            valid_prospect_ids = [
                prospect_id
                for prospect_id in prospect_ids
                if prospect_id not in excluded_ids
            ]

            if not valid_prospect_ids:
                conn.rollback()

                return jsonify({
                    "ok": False,
                    "error": (
                        "Todos los prospectos seleccionados están dados "
                        "de baja o bloqueados para envíos"
                    ),
                    "unsubscribed_ids": sorted(unsubscribed_ids),
                    "suppressed_ids": sorted(suppressed_ids),
                    "suppressed_reasons": suppressed_reasons
                }), 400

            print(f"PROSPECT IDS RECIBIDOS: {prospect_ids}")
            print(f"TOTAL RECIBIDOS: {len(prospect_ids)}")
            print(f"UNSUBSCRIBED IDS: {sorted(unsubscribed_ids)}")
            print(f"TOTAL UNSUBSCRIBED: {len(unsubscribed_ids)}")
            print(f"SUPPRESSED IDS: {sorted(suppressed_ids)}")
            print(f"TOTAL SUPPRESSED: {len(suppressed_ids)}")
            print(f"TOTAL VÁLIDOS: {len(valid_prospect_ids)}")

            # 2) Insertar cabecera solo si queda algún prospecto válido
            sql_target = """
                INSERT INTO prospect_targets (name)
                VALUES (%s)
            """

            cur.execute(sql_target, (nombre_target,))
            new_id = cur.lastrowid

            # 3) Insertar solo prospectos válidos
            sql_item = """
                INSERT INTO prospect_target_items (
                    target_id,
                    prospect_id
                )
                VALUES (%s, %s)
            """

            params_items = [
                (new_id, prospect_id)
                for prospect_id in valid_prospect_ids
            ]

            cur.executemany(sql_item, params_items)

        conn.commit()

        return jsonify({
            "ok": True,
            "id": new_id,
            "nombre_target": nombre_target,
            "total_seleccionados": len(prospect_ids),
            "total_prospectos": len(valid_prospect_ids),
            "total_omitidos": len(excluded_ids),
            "omitidos_unsubscribe": len(unsubscribed_ids),
            "omitidos_bloqueados": len(suppressed_ids),
            "unsubscribed_ids": sorted(unsubscribed_ids),
            "suppressed_ids": sorted(suppressed_ids),
            "suppressed_reasons": suppressed_reasons,
            "message": (
                f"Prospect target '{nombre_target}' almacenado con "
                f"{len(valid_prospect_ids)} prospectos. "
                f"Se omitieron {len(unsubscribed_ids)} dados de baja y "
                f"{len(suppressed_ids)} bloqueados."
            )
        }), 201

    except pymysql.err.IntegrityError as e:
        conn.rollback()

        errno = e.args[0] if e.args else None
        errmsg = e.args[1] if len(e.args) > 1 else str(e)

        print(
            f"DB IntegrityError {errno}: {errmsg} | "
            f"nombre_target={repr(nombre_target)}"
        )

        return jsonify({
            "ok": False,
            "error": f"MySQL {errno}: {errmsg}"
        }), 400

    except pymysql.err.Error as e:
        conn.rollback()

        print(f"DB Error: {e}")

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

    except Exception as e:
        conn.rollback()

        print(f"Error inesperado: {e}")

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

    finally:
        try:
            conn.close()
        except Exception:
            pass

@application.route("/bloquear_contacto", methods=["POST"])
def bloquear_contacto():
    data = request.get_json(silent=True) or request.form

    entity_kind = data.get("entity_kind")
    entity_id = data.get("entity_id")

    if not entity_kind or not entity_id:
        return jsonify({
            "ok": False,
            "message": "Faltan entity_kind o entity_id"
        }), 400

    try:
        entity_id = int(entity_id)
    except (TypeError, ValueError):
        return jsonify({
            "ok": False,
            "message": "entity_id no válido"
        }), 400

    if entity_kind == "lead":
        contacto = db.session.get(LeadForm, entity_id)

    elif entity_kind == "prospect":
        contacto = db.session.get(ProspectsIA, entity_id)

    else:
        return jsonify({
            "ok": False,
            "message": "Tipo de contacto no válido"
        }), 400

    if not contacto:
        return jsonify({
            "ok": False,
            "message": "Contacto no encontrado"
        }), 404

    contacto.email_suppressed = True
    contacto.email_suppressed_at = datetime.utcnow()
    contacto.email_suppressed_reason = "Bloqueado manualmente"

    try:
        db.session.commit()

        return jsonify({
            "ok": True,
            "message": "Contacto bloqueado correctamente",
            "entity_kind": entity_kind,
            "entity_id": entity_id
        })

    except Exception as exc:
        db.session.rollback()

        print("Error bloqueando contacto:", exc)

        return jsonify({
            "ok": False,
            "message": "No se pudo bloquear el contacto"
        }), 500



@application.route("/store_lead_targets", methods=["POST"])
def store_lead_targets():
    data = request.get_json(silent=True) or {}

    nombre_target = (data.get("nombre_target") or "").strip()
    lead_ids = data.get("lead_ids") or []

    if not nombre_target:
        return jsonify({
            "ok": False,
            "error": "nombre_target es obligatorio"
        }), 400

    if not isinstance(lead_ids, list) or not lead_ids:
        return jsonify({
            "ok": False,
            "error": "lead_ids debe ser una lista con al menos un id"
        }), 400

    try:
        lead_ids = [int(x) for x in lead_ids]
    except (TypeError, ValueError):
        return jsonify({
            "ok": False,
            "error": "Todos los lead_ids deben ser numéricos"
        }), 400

    # Quitar duplicados manteniendo el orden
    lead_ids = list(dict.fromkeys(lead_ids))

    print(
        f"[INFO] Almacenando lead target '{nombre_target}' "
        f"con {len(lead_ids)} leads: {lead_ids}"
    )

    creds = get_db_credentials("secretoBC/Mysql")
    dbname = "bc_pruebas" if BD == "PRUEBAS" else creds["dbname"]

    print(
        f"Conectando a la base de datos con host: {creds['host']}, "
        f"usuario: {creds['username']}, base de datos: {dbname}"
    )

    conn = pymysql.connect(
        host=creds["host"],
        user=creds["username"],
        password=creds["password"],
        database=dbname,
        port=int(creds.get("port", 3306)),
        autocommit=False
    )

    try:
        with conn.cursor() as cur:

            # 1. Validar que todos los leads existan y comprobar
            # baja voluntaria y bloqueo de email.
            placeholders = ",".join(["%s"] * len(lead_ids))

            sql_check = f"""
                SELECT
                    id,
                    COALESCE(unsubscribed, 0) AS unsubscribed,
                    COALESCE(email_suppressed, 0) AS email_suppressed,
                    email_suppressed_reason
                FROM lead_forms
                WHERE id IN ({placeholders})
            """

            cur.execute(sql_check, tuple(lead_ids))
            rows = cur.fetchall()

            existing_ids = {row[0] for row in rows}

            unsubscribed_ids = {
                row[0]
                for row in rows
                if int(row[1] or 0) == 1
            }

            suppressed_ids = {
                row[0]
                for row in rows
                if int(row[2] or 0) == 1
            }

            suppressed_reasons = {
                row[0]: row[3]
                for row in rows
                if int(row[2] or 0) == 1
            }

            missing_ids = [
                lead_id
                for lead_id in lead_ids
                if lead_id not in existing_ids
            ]

            if missing_ids:
                conn.rollback()

                return jsonify({
                    "ok": False,
                    "error": "Algunos leads no existen",
                    "missing_ids": missing_ids
                }), 400

            # Un lead puede estar dado de baja y bloqueado a la vez.
            excluded_ids = unsubscribed_ids | suppressed_ids

            valid_lead_ids = [
                lead_id
                for lead_id in lead_ids
                if lead_id not in excluded_ids
            ]

            if not valid_lead_ids:
                conn.rollback()

                return jsonify({
                    "ok": False,
                    "error": (
                        "Todos los leads seleccionados están dados de baja "
                        "o bloqueados para envíos"
                    ),
                    "unsubscribed_ids": sorted(unsubscribed_ids),
                    "suppressed_ids": sorted(suppressed_ids),
                    "suppressed_reasons": suppressed_reasons
                }), 400

            # 2. Crear la cabecera solo después de validar que
            # queda al menos un lead válido.
            sql_target = """
                INSERT INTO lead_targets (nombre_target)
                VALUES (%s)
            """

            cur.execute(sql_target, (nombre_target,))
            new_id = cur.lastrowid

            # 3. Insertar únicamente los leads válidos.
            sql_item = """
                INSERT INTO lead_target_items (target_id, lead_id)
                VALUES (%s, %s)
            """

            params_items = [
                (new_id, lead_id)
                for lead_id in valid_lead_ids
            ]

            cur.executemany(sql_item, params_items)

        conn.commit()

        return jsonify({
            "ok": True,
            "id": new_id,
            "nombre_target": nombre_target,
            "total_leads": len(valid_lead_ids),

            "omitidos_unsubscribe": len(unsubscribed_ids),
            "omitidos_bloqueados": len(suppressed_ids),
            "total_omitidos": len(excluded_ids),

            "message": (
                f"Lead target '{nombre_target}' almacenado con "
                f"{len(valid_lead_ids)} leads. "
                f"Se omitieron {len(excluded_ids)} contactos: "
                f"{len(unsubscribed_ids)} dados de baja y "
                f"{len(suppressed_ids)} bloqueados."
            )
        }), 201

    except pymysql.err.IntegrityError as e:
        conn.rollback()

        errno = e.args[0] if e.args else None
        errmsg = e.args[1] if len(e.args) > 1 else str(e)

        print(
            f"DB IntegrityError {errno}: {errmsg} | "
            f"nombre_target={repr(nombre_target)}"
        )

        return jsonify({
            "ok": False,
            "error": f"MySQL {errno}: {errmsg}"
        }), 400

    except pymysql.err.Error as e:
        conn.rollback()

        print(f"DB Error: {e}")

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

    except Exception as e:
        conn.rollback()

        print(f"Error inesperado: {e}")

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

    finally:
        try:
            conn.close()
        except Exception:
            pass

def clean_text(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text


def clean_int(value):
    if pd.isna(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None



@application.route("/consultar_prospectos_IA", methods=["GET", "POST"])
def consultar_prospectos_IA():

    estado = request.args.get("estado", "").strip()
    origen = request.args.get("origen", "").strip()

    # Contexto de navegación e informes
    modo = request.args.get("modo", "prospectos").strip()
    campaign_id = request.args.get("campaign_id", type=int)
    campaign_name = request.args.get("campaign_name", "").strip()

    modos_validos = (
        "prospectos",
        "aperturas",
        "clicks",
        "rebotes",
        "bajas"
    )

    if modo not in modos_validos:
        modo = "prospectos"

    titulos = {
        "prospectos": "Consultar prospectos",
        "aperturas": "Prospectos que han abierto la campaña: ",
        "clicks": "Prospectos que han hecho clic en la campaña: ",
        "rebotes": "Prospectos con rebote en la campaña:",
        "bajas": "Prospectos dados de baja en la campaña:",
    }

    titulo = titulos[modo]



    creds = get_db_credentials("secretoBC/Mysql")
    dbname = "bc_pruebas" if BD == "PRUEBAS" else creds["dbname"]

    conn = pymysql.connect(
        host=creds["host"],
        user=creds["username"],
        password=creds["password"],
        database=dbname,
        port=int(creds.get("port", 3306)),
        cursorclass=pymysql.cursors.DictCursor
    )

    try:
        condiciones_prospecto = []
        params_prospecto = []

        condiciones_campaign = []
        params_campaign = []

        # Siempre mostramos registros clasificados como prospect
        condiciones_prospecto.append("p.lead_status = %s")
        params_prospecto.append("prospect")

        # Filtros normales de prospectos
        if estado and estado != "Todos":
            condiciones_prospecto.append("p.estado = %s")
            params_prospecto.append(estado)

        if origen:
            condiciones_prospecto.append("ps.description = %s")
            params_prospecto.append(origen)

        # Filtro para bajas
        if modo == "bajas":
            condiciones_prospecto.append("p.unsubscribed = 1")

        # Informes asociados a una campaña
        if campaign_id:
            condiciones_campaign.append("cr.campaign_id = %s")
            params_campaign.append(campaign_id)

        if modo == "aperturas":
            condiciones_campaign.append("cr.opened_at IS NOT NULL")

        elif modo == "clicks":
            condiciones_campaign.append("cr.clicked_at IS NOT NULL")

        elif modo == "rebotes":
            condiciones_campaign.append("cr.bounced_at IS NOT NULL")

        # En los informes necesitamos que el prospecto pertenezca
        # obligatoriamente a la campaña
        if modo in ("aperturas", "clicks", "rebotes", "bajas"):
            tipo_join = "INNER JOIN"
        else:
            tipo_join = "LEFT JOIN"

        condiciones = condiciones_prospecto + condiciones_campaign
        params = params_prospecto + params_campaign

        where_sql = ""

        if condiciones:
            where_sql = "WHERE " + " AND ".join(condiciones)

        with conn.cursor() as cur:

            sql = f"""
                SELECT
                    p.id,
                    p.fecha,
                    p.idioma,
                    p.pais,
                    p.email,
                    p.club,
                    p.estado,
                    p.tipo,
                    p.propietario,
                    p.num_pistas,
                    p.web,
                    p.youtube,
                    p.instagram,
                    p.linkedin_club,
                    p.linkedin_propietario,
                    p.booking_app,
                    p.proveedor_pistas,
                    p.unsubscribed,
                    p.unsubscribed_at,
                    ps.description AS origen,

                    COUNT(cr.id) AS total_campanas,

                    MAX(cr.opened_at) AS ultima_apertura,
                    MAX(cr.clicked_at) AS ultimo_click,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN cr.opened_at IS NOT NULL THEN 1
                                ELSE 0
                            END
                        ),
                        0
                    ) AS total_abiertas,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN cr.clicked_at IS NOT NULL THEN 1
                                ELSE 0
                            END
                        ),
                        0
                    ) AS total_campanas_clicadas,

                    COALESCE(
                        SUM(COALESCE(cr.click_count, 0)),
                        0
                    ) AS total_clicks

                FROM prospects_IA p

                {tipo_join} campaign_recipients cr
                    ON cr.entity_id = p.id
                AND cr.entity_kind = 'prospect'

                LEFT JOIN prospect_sources ps
                    ON ps.id = p.source_id

                {where_sql}

                GROUP BY
                    p.id,
                    p.fecha,
                    p.idioma,
                    p.pais,
                    p.email,
                    p.club,
                    p.estado,
                    p.tipo,
                    p.propietario,
                    p.num_pistas,
                    p.web,
                    p.youtube,
                    p.instagram,
                    p.linkedin_club,
                    p.linkedin_propietario,
                    p.booking_app,
                    p.proveedor_pistas,
                    p.unsubscribed,
                    p.unsubscribed_at,
                    ps.description

                ORDER BY
                    p.fecha DESC,
                    p.id DESC
            """

            print("[DEBUG] modo:", modo)
            print("[DEBUG] campaign_id:", campaign_id)
            print("[DEBUG] SQL:", sql)
            print("[DEBUG] params:", params)

            cur.execute(sql, params)
            rows = cur.fetchall()

            cur.execute("""
                SELECT DISTINCT
                    ps.description AS origen
                FROM prospects_IA p
                LEFT JOIN prospect_sources ps
                    ON ps.id = p.source_id
                WHERE ps.description IS NOT NULL
                  AND TRIM(ps.description) <> ''
                ORDER BY ps.description
            """)

            origenes_rows = cur.fetchall()

        origenes = [row["origen"] for row in origenes_rows]

    finally:
        conn.close()

    titulos = {
        "prospectos": "Consultar prospectos",
        "aperturas": "Prospectos que abrieron la campaña",
        "clicks": "Prospectos que hicieron clic en la campaña :",
        "rebotes": "Prospectos con rebote en la campaña :",
        "bajas": "Prospectos dados de bajaen la campaña :"
    }

    titulo = titulos.get(modo, "Consultar prospectos")

    return render_template(
        "consultar_prospectos_IA.html",
        prospectos=rows,
        estado=estado,
        origen=origen,
        origenes=origenes,
        modo=modo,
        campaign_id=campaign_id,
        campaign_name=campaign_name,
        titulo=titulo,
        entity_kind="prospect"
    )



@application.route("/import_bd_IA", methods=["GET", "POST"])
def import_bd_IA():
    if request.method == "GET":
        return render_template("import_bd_IA.html")

    file = request.files.get("file")
    description = request.form.get("description", "").strip()

    if not file or file.filename == "":
        flash("Debes seleccionar un archivo Excel.", "error")
        return redirect(url_for("import_bd_IA"))

    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        flash("El archivo debe ser un Excel válido (.xlsx o .xls).", "error")
        return redirect(url_for("import_bd_IA"))

    file_name = file.filename.strip()
    default_description = file_name.rsplit(".", 1)[0] if "." in file_name else file_name
    if not description:
        description = default_description

    try:
        df = pd.read_excel(file)

        df.columns = [
            unidecode(str(col))
                .strip()
                .lower()
                .replace("\n", " ")
                .replace("\r", " ")
            for col in df.columns
        ]


        rename_map = {
            "fecha": "fecha",
            "idioma": "idioma",
            "país": "pais",
            "pais": "pais",
            "email": "email",
            "club": "club",
            "nombre club": "club",
            "nombre_club": "club",
            "estado": "estado",
            "propietario": "propietario",
            "nº pistas": "num_pistas",
            "n° pistas": "num_pistas",
            "num pistas": "num_pistas",
            "num_pistas": "num_pistas",
            "tipo": "tipo",
            "web": "web",
            "youtube": "youtube",
            "instagram": "instagram",
            "linkedin club": "linkedin_club",
            "linkedin_club": "linkedin_club",
            "linkedin propietario": "linkedin_propietario",
            "linkedin_propietario": "linkedin_propietario",
            "booking app": "booking_app",
            "booking_app": "booking_app",
            "proveedor pistas": "proveedor_pistas",
            "proveedor_pistas": "proveedor_pistas",
        }

        df = df.rename(columns=rename_map)

        required_columns = {"fecha", "idioma", "pais", "email", "estado"}
        missing = required_columns - set(df.columns)

        if missing:
            flash(
                f"Faltan columnas obligatorias: {', '.join(sorted(missing))}",
                "error"
            )
            return redirect(url_for("import_bd_IA"))

        estados_validos = {"operativo", "renovacion", "concepto", "proyecto"}

        creds = get_db_credentials("secretoBC/Mysql")
        dbname = "bc_pruebas" if (BD == "PRUEBAS") else creds["dbname"]

        print(
            f"Conectando a la base de datos con host: {creds['host']}, "
            f"usuario: {creds['username']}, base de datos: {dbname}"
        )

        conn = pymysql.connect(
            host=creds["host"],
            user=creds["username"],
            password=creds["password"],
            database=dbname,
            port=int(creds.get("port", 3306)),
            autocommit=False,
            cursorclass=pymysql.cursors.Cursor
        )

        procesadas = 0
        omitidas = 0
        errores = []

        sql_insert_source = """
            INSERT INTO prospect_sources (
                file_name,
                description,
                created_at
            ) VALUES (
                %s, %s, NOW()
            )
        """

        sql = """
            INSERT INTO prospects_IA (
                fecha,
                idioma,
                pais,
                email,
                club,
                estado,
                propietario,
                num_pistas,
                tipo,
                web,
                youtube,
                instagram,
                linkedin_club,
                linkedin_propietario,
                booking_app,
                proveedor_pistas,
                unsubscribed,
                unsubscribed_at,
                source_id
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            ON DUPLICATE KEY UPDATE
                fecha = VALUES(fecha),
                idioma = VALUES(idioma),
                pais = VALUES(pais),
                club = VALUES(club),
                estado = VALUES(estado),
                propietario = VALUES(propietario),
                num_pistas = VALUES(num_pistas),
                tipo = VALUES(tipo),
                web = VALUES(web),
                youtube = VALUES(youtube),
                instagram = VALUES(instagram),
                linkedin_club = VALUES(linkedin_club),
                linkedin_propietario = VALUES(linkedin_propietario),
                booking_app = VALUES(booking_app),
                proveedor_pistas = VALUES(proveedor_pistas),
                source_id = VALUES(source_id)
        """

        try:
            with conn.cursor() as cur:
                cur.execute(sql_insert_source, (file_name, description))
                source_id = cur.lastrowid

                for index, row in df.iterrows():
                    try:
                        email = clean_text(row.get("email"))
                        if not email:
                            errores.append(f"Fila {index + 2}: email vacío")
                            omitidas += 1
                            continue
                        email = email.lower()

                        if lead_exists_for_prospect(cur, email):
                            errores.append(f"Fila {index + 2}: omitida porque ya existe como lead ({email})")
                            omitidas += 1
                            continue

                        estado = clean_text(row.get("estado"))
                        if not estado:
                            errores.append(f"Fila {index + 2}: estado vacío")
                            omitidas += 1
                            continue
                        estado = estado.lower()

                        if estado not in estados_validos:
                            errores.append(f"Fila {index + 2}: estado inválido '{estado}'")
                            omitidas += 1
                            continue

                        fecha = pd.to_datetime(row.get("fecha"), errors="coerce")
                        if pd.isna(fecha):
                            errores.append(f"Fila {index + 2}: fecha inválida")
                            omitidas += 1
                            continue

                        idioma = clean_text(row.get("idioma"))
                        pais = clean_text(row.get("pais"))

                        if not idioma:
                            errores.append(f"Fila {index + 2}: idioma vacío")
                            omitidas += 1
                            continue

                        if not pais:
                            errores.append(f"Fila {index + 2}: país vacío")
                            omitidas += 1
                            continue

                        values = (
                            fecha.date(),
                            idioma,
                            pais,
                            email,
                            clean_text(row.get("club")),
                            estado,
                            clean_text(row.get("propietario")),
                            clean_int(row.get("num_pistas")),
                            clean_text(row.get("tipo")),
                            clean_text(row.get("web")),
                            clean_text(row.get("youtube")),
                            clean_text(row.get("instagram")),
                            clean_text(row.get("linkedin_club")),
                            clean_text(row.get("linkedin_propietario")),
                            clean_text(row.get("booking_app")),
                            clean_text(row.get("proveedor_pistas")),
                            False,
                            None,
                            source_id
                        )

                        cur.execute(sql, values)
                        procesadas += 1

                    except Exception as e:
                        errores.append(f"Fila {index + 2}: {str(e)}")
                        omitidas += 1

                conn.commit()

        except Exception as e:
            conn.rollback()
            flash(f"Error al importar en base de datos: {str(e)}", "error")
            return redirect(url_for("import_bd_IA"))

        finally:
            conn.close()

        flash(
            f"Importación completada. Filas procesadas: {procesadas}. Omitidas: {omitidas}.",
            "success"
        )

        if errores:
            flash("Errores detectados:<br>" + "<br>".join(errores[:10]), "warning")

        return redirect(url_for("import_bd_IA"))

    except Exception as e:
        flash(f"Error al leer el archivo: {str(e)}", "error")
        return redirect(url_for("import_bd_IA"))
    




@application.route("/exportar_datos_campanas", methods=["GET", "POST"])
def exportar_datos_campanas():

    if request.method == "GET":
        return render_template("exportacion_datos_campanas.html")
    
    connection = None

    try:
        creds = get_db_credentials("secretoBC/Mysql")
        BD = request.args.get("BD", "")
        dbname = "bc_pruebas" if BD == "PRUEBAS" else creds["dbname"]

        connection = pymysql.connect(
            host=creds["host"],
            user=creds["username"],
            password=creds["password"],
            database=dbname,
            port=int(creds.get("port", 3306)),
            cursorclass=pymysql.cursors.DictCursor,
            autocommit=True,
        )

        query = """
            SELECT
                c.id AS campaign_id,
                c.name AS campaign_name,
                c.campaign_type,
                c.status AS campaign_status,
                c.created_at AS campaign_created_at,
                c.scheduled_at,
                c.sent_at AS campaign_sent_at,

                cr.id AS recipient_id,
                cr.email,
                cr.entity_kind,
                cr.entity_id,

                cr.pais,
                cr.idioma,
                cr.origen,
                cr.segment,
                cr.tipo_lead,
                cr.estado AS recipient_estado,

                cr.send_status,
                cr.sent_at,
                cr.delivered_at,
                cr.opened_at,
                cr.clicked_at,
                cr.bounced_at,
                cr.complained_at,
                cr.last_event,
                cr.click_count,
                cr.tracking_id,
                cr.ses_message_id,

                p.id AS prospect_id,
                p.fecha AS prospect_created_date,
                p.club,
                p.tipo AS prospect_tipo,
                p.estado AS prospect_estado,
                p.propietario,
                p.num_pistas,
                p.web,
                p.lead_status,
                p.lead_converted_at,

                lf.id AS lead_id,
                lf.name AS lead_name,
                lf.created_at AS lead_created_at,
                lf.quote_number,
                lf.estado AS lead_estado,
                lf.probabilidad_exito,
                lf.cantidad_total,
                lf.descuento_total,

                CASE WHEN lf.proforma_solicitada = 1 THEN 'Sí' ELSE 'No' END AS solicito_proforma,
                lf.fecha_solicitud_proforma,

                CASE WHEN lf.renting_solicitado = 1 THEN 'Sí' ELSE 'No' END AS solicito_renting,
                lf.fecha_solicitud_renting,

                CASE WHEN lf.renting_concedido = 1 THEN 'Sí' ELSE 'No' END AS renting_aprobado,
                lf.fecha_concedido_renting,

                CASE WHEN lf.renting_denegado = 1 THEN 'Sí' ELSE 'No' END AS renting_denegado,
                lf.fecha_denegado_renting,

                CASE WHEN p.unsubscribed = 1 THEN 'Sí' ELSE 'No' END AS prospect_unsubscribed,
                p.unsubscribed_at AS prospect_unsubscribed_at,

                CASE WHEN lf.unsubscribed = 1 THEN 'Sí' ELSE 'No' END AS lead_unsubscribed,
                lf.unsubscribed_at AS lead_unsubscribed_at

            FROM campaign_recipients cr
            INNER JOIN campaigns c ON c.id = cr.campaign_id
            LEFT JOIN prospects_IA p
                ON cr.entity_kind = 'prospect'
            AND cr.entity_id = p.id
            LEFT JOIN (
                SELECT lf1.*
                FROM lead_forms lf1
                INNER JOIN (
                    SELECT email, MAX(id) AS ultimo_id
                    FROM lead_forms
                    GROUP BY email
                ) ult ON ult.ultimo_id = lf1.id
            ) lf ON lf.email = cr.email
            ORDER BY c.id, cr.sent_at, cr.email;
            """

        with connection.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()

        df = pd.DataFrame(rows)

        df = df.rename(columns={
            "campaign_id": "ID Campaña",
            "campaign_name": "Campaña",
            "campaign_type": "Tipo campaña",
            "campaign_status": "Estado campaña",
            "campaign_created_at": "Fecha creación campaña",
            "scheduled_at": "Fecha programada",
            "campaign_sent_at": "Fecha envío campaña",
            "recipient_id": "ID Destinatario",
            "email": "Email",
            "entity_kind": "Tipo origen",
            "entity_id": "ID origen",
            "pais": "País",
            "idioma": "Idioma",
            "origen": "Origen",
            "segment": "Segmento",
            "tipo_lead": "Tipo lead",
            "recipient_estado": "Estado contacto",
            "send_status": "Estado envío",
            "sent_at": "Fecha envío email",
            "delivered_at": "Fecha entrega",
            "opened_at": "Fecha apertura",
            "clicked_at": "Fecha clic",
            "bounced_at": "Fecha rebote",
            "complained_at": "Fecha queja",
            "last_event": "Último evento",
            "click_count": "Número de clics",
            "tracking_id": "Tracking ID",
            "ses_message_id": "ID mensaje SES",
            "prospect_id": "ID Prospecto",
            "prospect_created_date": "Fecha alta prospecto",
            "club": "Club",
            "prospect_tipo": "Tipo prospecto",
            "prospect_estado": "Estado prospecto",
            "propietario": "Comercial asignado",
            "num_pistas": "Número de pistas",
            "web": "Web",
            "lead_status": "Estado comercial",
            "lead_converted_at": "Fecha conversión a lead",
            "lead_id": "ID Lead",
            "lead_name": "Nombre lead",
            "lead_created_at": "Fecha creación oferta",
            "quote_number": "Número oferta",
            "lead_estado": "Estado oferta",
            "probabilidad_exito": "Probabilidad de éxito",
            "cantidad_total": "Importe oferta",
            "descuento_total": "Descuento total",
            "solicito_proforma": "Solicitó proforma",
            "fecha_solicitud_proforma": "Fecha solicitud proforma",
            "solicito_renting": "Solicitó renting",
            "fecha_solicitud_renting": "Fecha solicitud renting",
            "renting_aprobado": "Renting aprobado",
            "fecha_concedido_renting": "Fecha aprobación renting",
            "renting_denegado": "Renting denegado",
            "fecha_denegado_renting": "Fecha denegación renting",
            "prospect_unsubscribed": "Baja prospecto",
            "prospect_unsubscribed_at": "Fecha baja prospecto",
            "lead_unsubscribed": "Baja lead",
            "lead_unsubscribed_at": "Fecha baja lead",
        })


       
       

        print(df.head(5).to_string())
        print(df.dtypes)

        

        output = BytesIO()




        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # Escribir el DataFrame
            df.to_excel(writer, index=False, sheet_name="Marketing completo")

            ws = writer.book["Marketing completo"]

            # Congelar primera fila
            ws.freeze_panes = "A2"

            # Activar filtros
            ws.auto_filter.ref = ws.dimensions

            # ====== FORMATO DEL ENCABEZADO ======
            from openpyxl.styles import Font, PatternFill

            header_fill = PatternFill(
                fill_type="solid",
                start_color="1F4E78",
                end_color="1F4E78"
            )

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill

            # ====== AJUSTAR ANCHO COLUMNAS ======
            for column in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 40)



        output.seek(0)

        document_no = f"Datos_Campañas_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

        dropbox_path = f"/2026 PPT - Marketing/Datos Campañas/{document_no}"




        ruta = subir_a_dropbox(
            output.getvalue(),
            dropbox_path
        )

        print(f"Archivo guardado en {ruta}")

        







       
    except Exception as e:
        print("ERROR exportando datos campañas:", e)
        return {"error": str(e)}, 500

    finally:
        if connection:
            connection.close()


    return render_template(
        "exportacion_datos_campanas.html",
        ok=True,
        archivo=document_no,
        ruta=ruta
    )

@application.route('/base', methods=['GET', 'POST'])
def base():
    
    return render_template('base.html')


@application.route('/logout', methods=['GET', 'POST'])
def logout():
    session.clear()  # Elimina todos los datos de sesión
    return redirect(url_for('login'))  # Cambiá 'login' por tu vista de inicio o login

    


# Función para crear la base de datos si no existe
#def crear_base_si_no_existe():
#    with application.app_context():
#        db.create_all()

# Crear la base de datos y las tablas
crear_base_si_no_existe()

# Determinar el entorno de ejecución



if  __name__ == "__main__":     
    
    
    
    application.run(host='0.0.0.0', port=8000, debug=True, use_reloader=False)




